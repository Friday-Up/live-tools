"""测试结果报告生成。"""
from __future__ import annotations

import tempfile
import unittest
from datetime import datetime

from openpyxl import load_workbook

from room_creator.models import BatchResult, RoomCreateResult
from room_creator.report_writer import write_batch_report


class TestWriteBatchReport(unittest.TestCase):
    def test_writes_result_excel(self):
        result = BatchResult()
        result.results = [
            RoomCreateResult(
                row_index=2,
                title="直播间一号",
                start_time=datetime(2026, 7, 1, 20, 0, 0),
                live_form="正式直播",
                live_direction="竖屏",
                live_location="不显示地点",
                live_category="多品类",
                success=True,
            ),
            RoomCreateResult(
                row_index=3,
                title="直播间二号",
                start_time=datetime(2026, 7, 2, 20, 0, 0),
                live_form="正式直播",
                live_direction="竖屏",
                live_location="不显示地点",
                live_category="多品类",
                success=False,
                error="标题重复",
            ),
        ]
        result.created_count = 1
        result.failed_count = 1

        with tempfile.TemporaryDirectory() as tmpdir:
            path = write_batch_report(result, tmpdir)
            self.assertTrue(path.exists())
            wb = load_workbook(path)
            ws = wb.active
            self.assertIsNotNone(ws)
            self.assertEqual(ws["A1"].value, "原始行号")
            self.assertEqual(ws["H2"].value, "成功")
            self.assertEqual(ws["H3"].value, "失败")
            self.assertEqual(ws["I3"].value, "标题重复")


if __name__ == "__main__":
    unittest.main()
