"""读取业务输入 Excel 并解析为直播间创建行。"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from . import config
from .models import RoomCreateRow


@dataclass
class ColumnMapping:
    """Excel 列映射结果。"""

    title_col: Optional[str] = None
    cover_col: Optional[str] = None
    start_time_col: Optional[str] = None
    live_form_col: Optional[str] = None
    live_direction_col: Optional[str] = None
    live_location_col: Optional[str] = None
    live_category_col: Optional[str] = None

    def is_valid(self) -> bool:
        return bool(self.title_col and self.start_time_col)


def _normalize_header(header: str) -> str:
    return str(header).strip().replace(" ", "").replace("（", "").replace("）", "").replace("*", "")


def _find_column(headers: list[str], aliases: list[str]) -> Optional[str]:
    """根据别名列表找到匹配的列名（优先精确匹配，再做包含匹配）。"""
    normalized = [(_normalize_header(h), h) for h in headers]
    normalized_aliases = [_normalize_header(a) for a in aliases]

    # 精确匹配
    for norm, original in normalized:
        if norm in normalized_aliases:
            return original

    # 包含匹配：列名包含某个别名
    for norm, original in normalized:
        for alias in normalized_aliases:
            if alias in norm or norm in alias:
                return original

    return None


def inspect_workbook(file_path: str | Path) -> tuple[ColumnMapping, list[str]]:
    """检查 Excel 表头并推荐列映射。"""
    file_path = Path(file_path)
    wb = load_workbook(file_path, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            raise ValueError("Excel 中没有工作表")

        headers = [str(cell.value or "").strip() for cell in ws[1]]
    finally:
        wb.close()

    mapping = ColumnMapping(
        title_col=_find_column(headers, config.TITLE_COLUMN_ALIASES),
        cover_col=_find_column(headers, config.COVER_COLUMN_ALIASES),
        start_time_col=_find_column(headers, config.START_TIME_COLUMN_ALIASES),
        live_form_col=_find_column(headers, config.FORM_COLUMN_ALIASES),
        live_direction_col=_find_column(headers, config.DIRECTION_COLUMN_ALIASES),
        live_location_col=_find_column(headers, config.LOCATION_COLUMN_ALIASES),
        live_category_col=_find_column(headers, config.CATEGORY_COLUMN_ALIASES),
    )
    return mapping, headers


def _parse_datetime(value) -> datetime:
    """解析开播时间，支持多种格式。"""
    if value is None:
        raise ValueError("开播时间为空")

    if isinstance(value, datetime):
        return value

    text = str(value).strip()
    if not text:
        raise ValueError("开播时间为空")

    for fmt in config.DATETIME_INPUT_FORMATS:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue

    raise ValueError(f"开播时间格式无法识别: {text}")


def read_room_rows(
    file_path: str | Path,
    mapping: ColumnMapping,
) -> list[RoomCreateRow]:
    """根据列映射读取所有创建行。"""
    file_path = Path(file_path)
    if not mapping.is_valid():
        raise ValueError("列映射不完整：必须包含直播标题和开播时间列")

    wb = load_workbook(file_path, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            raise ValueError("Excel 中没有工作表")

        rows: list[RoomCreateRow] = []
        header_values = [str(cell.value or "").strip() for cell in ws[1]]

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            def cell_value(col_name: Optional[str]):
                if not col_name:
                    return None
                try:
                    idx = header_values.index(col_name)
                    return row[idx] if idx < len(row) else None
                except ValueError:
                    return None

            title = cell_value(mapping.title_col)
            start_time = cell_value(mapping.start_time_col)

            # 标题和时间为空则跳过（视为空行）
            if title is None or str(title).strip() == "":
                continue
            if start_time is None or str(start_time).strip() == "":
                raise ValueError(f"第 {row_idx} 行开播时间为空")

            try:
                parsed_time = _parse_datetime(start_time)
            except ValueError as exc:
                raise ValueError(f"第 {row_idx} 行开播时间格式错误：{exc}")

            live_category = str(cell_value(mapping.live_category_col) or config.DEFAULT_LIVE_CATEGORY).strip()
            live_category = config.CATEGORY_OPTION_MAP.get(live_category, live_category)

            rows.append(
                RoomCreateRow(
                    row_index=row_idx,
                    title=str(title).strip(),
                    start_time=parsed_time,
                    cover=str(cell_value(mapping.cover_col) or "").strip(),
                    live_form=str(cell_value(mapping.live_form_col) or config.DEFAULT_LIVE_FORM).strip(),
                    live_direction=str(cell_value(mapping.live_direction_col) or config.DEFAULT_LIVE_DIRECTION).strip(),
                    live_location=str(cell_value(mapping.live_location_col) or config.DEFAULT_LIVE_LOCATION).strip(),
                    live_category=live_category,
                )
            )

        return rows
    finally:
        wb.close()
