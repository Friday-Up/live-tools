"""生成批量创建结果报告 Excel。"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from . import config
from .models import BatchResult


def write_batch_report(result: BatchResult, output_dir: str | Path) -> Path:
    """把批次结果写入 Excel，返回文件路径。"""
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    output_path = output_dir / f"直播间创建结果_{timestamp}_.xlsx"

    wb = Workbook()
    try:
        ws = wb.active
        if ws is None:
            raise RuntimeError("无法创建工作表")
        ws.title = "创建结果"

        headers = config.RESULT_COLUMNS
        ws.append(headers)

        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        for room_result in sorted(result.results, key=lambda r: r.row_index):
            ws.append(
                [
                    room_result.row_index,
                    room_result.title,
                    room_result.start_time.strftime("%Y-%m-%d %H:%M:%S"),
                    room_result.live_form,
                    room_result.live_direction,
                    room_result.live_location,
                    room_result.live_category,
                    "成功" if room_result.success else "失败",
                    room_result.error or ("" if room_result.success else "未知错误"),
                ]
            )

        # 汇总行
        ws.append([])
        ws.append(["汇总"])
        ws.append(["成功", result.created_count])
        ws.append(["失败", result.failed_count])
        ws.append(["跳过", result.skipped_count])
        if result.stopped_by_limit:
            ws.append(["终止原因", "超出每日创建上限"])
        if result.stopped_by_user:
            ws.append(["终止原因", "用户手动停止"])

        wb.save(output_path)
        return output_path
    finally:
        wb.close()
