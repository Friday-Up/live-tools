import io
import os
import tempfile
import time
import unittest
from datetime import date, timedelta
from pathlib import Path
from urllib.parse import unquote
from zipfile import ZipFile

from openpyxl import Workbook
from unittest.mock import patch

import app as web_app

from app import (
    create_app,
    resolve_live_dir,
    resolve_web_root,
    _format_price_diagnostics,
    _unlink_with_retries,
    parse_sku_input,
)


class PromotionBindingRoutesTest(unittest.TestCase):
    def test_resolves_source_and_packaged_runtime_paths(self):
        source_app = Path("/repo/live/live-web/app.py")
        packaged_exe = Path("C:/tools/Live-Tools-Web/Live-Tools-Web.exe")

        self.assertEqual(resolve_live_dir(file_path=source_app, frozen=False), Path("/repo/live"))
        self.assertEqual(resolve_live_dir(executable_path=packaged_exe, frozen=True), Path("C:/tools/Live-Tools-Web"))

        source_root = Path(tempfile.mkdtemp())
        web_root = source_root / "live-web"
        web_root.mkdir()
        self.assertEqual(resolve_web_root(source_root, source_app), web_root)

    def test_web_entry_forces_utf8_stdio_for_windows_log_redirection(self):
        source = Path("app.py").read_text(encoding="utf-8")

        self.assertIn('sys.stdout.reconfigure(encoding="utf-8", errors="replace")', source)
        self.assertIn('sys.stderr.reconfigure(encoding="utf-8", errors="replace")', source)

    def test_price_diagnostics_log_format_includes_source_counts(self):
        diagnostics = {
            "duration_ms": 12345,
            "spec_count": 24,
            "price_source_counts": {
                "ware-business": 20,
                "dom-fallback": 3,
                "dom": 2,
                "selected-dom": 1,
            },
        }

        self.assertEqual(
            _format_price_diagnostics(diagnostics),
            "诊断: 耗时 12.3s，规格 24，取价 ware=20/dom=5/selected=1",
        )

    def test_generate_promotion_binding_files(self):
        base_dir = Path(tempfile.mkdtemp())
        app = create_app(base_dir=base_dir)
        client = app.test_client()
        old_runtime_file = base_dir / "runtime" / "output" / "promotion-binding" / "old" / "old.xlsx"
        old_runtime_file.parent.mkdir(parents=True, exist_ok=True)
        old_runtime_file.write_bytes(b"old")
        old_time = time.time() - 8 * 24 * 60 * 60
        os.utime(old_runtime_file, (old_time, old_time))

        data = {
            "file": (
                io.BytesIO(self._business_workbook_bytes()),
                "业务提报.xlsx",
            )
        }

        response = client.post(
            "/api/promotion-binding/generate",
            data=data,
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["summary"]["success_count"], 2)
        self.assertIn("template_download_url", payload)
        self.assertIn("report_download_url", payload)

        template_response = client.get(payload["template_download_url"])
        report_response = client.get(payload["report_download_url"])

        self.assertEqual(template_response.status_code, 200)
        self.assertEqual(report_response.status_code, 200)
        template_disposition = unquote(template_response.headers["Content-Disposition"])
        report_disposition = unquote(report_response.headers["Content-Disposition"])
        self.assertIn("京东绑券上传模板_", template_disposition)
        self.assertIn("异常报告_", report_disposition)
        self.assertNotIn(payload["task_id"], template_disposition)
        self.assertNotIn(payload["task_id"], report_disposition)
        self.assertFalse(old_runtime_file.exists())
        self.assertTrue(str(app.config["PROMOTION_INPUT_DIR"]).startswith(str(base_dir / "runtime")))
        self.assertTrue(str(app.config["PROMOTION_OUTPUT_DIR"]).startswith(str(base_dir / "runtime")))
        self.assertTrue(app.config["PROMOTION_RESULTS"][payload["task_id"]]["template"].name.startswith("京东绑券上传模板_"))
        self.assertTrue(app.config["PROMOTION_RESULTS"][payload["task_id"]]["report"].name.startswith("异常报告_"))
        self.assertGreater(len(template_response.data), 1000)
        self.assertGreater(len(report_response.data), 1000)
        template_response.close()
        report_response.close()

    def test_preview_promotion_binding_columns(self):
        app = create_app(base_dir=Path(tempfile.mkdtemp()))
        client = app.test_client()

        response = client.post(
            "/api/promotion-binding/preview",
            data={"file": (io.BytesIO(self._business_workbook_bytes()), "业务提报.xlsx")},
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertTrue(payload["success"])
        self.assertTrue(payload["task_id"])
        self.assertEqual(payload["suggested_mapping"]["sku_col"], 1)
        self.assertEqual(payload["suggested_mapping"]["product_name_col"], 2)
        self.assertEqual(payload["suggested_mapping"]["code_col"], 3)
        self.assertEqual(payload["columns"][0]["header"], "skuID")
        self.assertEqual(payload["columns"][0]["sample_values"], ["1001", "1002"])
        self.assertTrue(Path(payload["path"]).exists())

    def test_preview_suggests_selling_point_column(self):
        app = create_app(base_dir=Path(tempfile.mkdtemp()))
        client = app.test_client()

        response = client.post(
            "/api/promotion-binding/preview",
            data={"file": (io.BytesIO(self._business_workbook_with_selling_point_bytes()), "业务提报.xlsx")},
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["suggested_mapping"]["selling_point_col"], 4)

    def test_generate_with_enable_selling_point(self):
        base_dir = Path(tempfile.mkdtemp())
        app = create_app(base_dir=base_dir)
        client = app.test_client()

        preview_response = client.post(
            "/api/promotion-binding/preview",
            data={"file": (io.BytesIO(self._business_workbook_with_selling_point_bytes()), "业务提报.xlsx")},
            content_type="multipart/form-data",
        )
        preview_payload = preview_response.get_json()
        self.assertTrue(preview_payload["success"])

        response = client.post(
            "/api/promotion-binding/generate",
            json={
                "task_id": preview_payload["task_id"],
                "column_mapping": {
                    "sku_col": 1,
                    "code_col": 3,
                    "selling_point_col": 4,
                },
                "enable_selling_point": True,
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["summary"]["selling_point_count"], 2)
        self.assertEqual(payload["messages"], [])

    def test_generate_warns_when_selling_point_column_not_found(self):
        base_dir = Path(tempfile.mkdtemp())
        app = create_app(base_dir=base_dir)
        client = app.test_client()

        preview_response = client.post(
            "/api/promotion-binding/preview",
            data={"file": (io.BytesIO(self._business_workbook_bytes()), "业务提报.xlsx")},
            content_type="multipart/form-data",
        )
        preview_payload = preview_response.get_json()
        self.assertTrue(preview_payload["success"])

        response = client.post(
            "/api/promotion-binding/generate",
            json={
                "task_id": preview_payload["task_id"],
                "column_mapping": {
                    "sku_col": 1,
                    "code_col": 3,
                },
                "enable_selling_point": True,
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertTrue(payload["success"])
        self.assertIn("未识别到短卖点列", payload["messages"])

    def test_generate_promotion_binding_files_with_selected_columns(self):
        base_dir = Path(tempfile.mkdtemp())
        app = create_app(base_dir=base_dir)
        client = app.test_client()

        preview_response = client.post(
            "/api/promotion-binding/preview",
            data={"file": (io.BytesIO(self._manual_mapping_workbook_bytes()), "业务提报.xlsx")},
            content_type="multipart/form-data",
        )
        preview_payload = preview_response.get_json()
        self.assertTrue(preview_payload["success"])
        self.assertIsNone(preview_payload["suggested_mapping"]["sku_col"])
        self.assertIsNone(preview_payload["suggested_mapping"]["code_col"])

        response = client.post(
            "/api/promotion-binding/generate",
            json={
                "task_id": preview_payload["task_id"],
                "column_mapping": {
                    "sku_col": 2,
                    "code_col": 4,
                },
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["summary"]["success_count"], 2)
        self.assertIn(payload["task_id"], app.config["PROMOTION_RESULTS"])

    def test_create_app_cleans_old_runtime_and_legacy_files(self):
        base_dir = Path(tempfile.mkdtemp())
        old_runtime_file = base_dir / "runtime" / "input" / "promotion-binding" / "old.xlsx"
        fresh_runtime_file = base_dir / "runtime" / "input" / "promotion-binding" / "fresh.xlsx"
        old_legacy_file = base_dir / "input" / "promotion-binding" / "legacy.xlsx"
        for file_path in [old_runtime_file, fresh_runtime_file, old_legacy_file]:
            file_path.parent.mkdir(parents=True, exist_ok=True)
            file_path.write_bytes(b"x")

        now = time.time()
        old_time = now - 3 * 24 * 60 * 60
        fresh_time = now - 24 * 60 * 60
        os.utime(old_runtime_file, (old_time, old_time))
        os.utime(old_legacy_file, (old_time, old_time))
        os.utime(fresh_runtime_file, (fresh_time, fresh_time))

        app = create_app(base_dir=base_dir)

        self.assertEqual(app.config["RUNTIME_RETENTION_DAYS"], 2)
        self.assertFalse(old_runtime_file.exists())
        self.assertFalse(old_legacy_file.exists())
        self.assertTrue(fresh_runtime_file.exists())
        self.assertEqual(app.config["PRICE_INPUT_DIR"], base_dir / "runtime" / "input" / "price-audit")

    def test_rejects_non_xlsx_upload(self):
        app = create_app(base_dir=Path(tempfile.mkdtemp()))
        client = app.test_client()

        response = client.post(
            "/api/promotion-binding/generate",
            data={"file": (io.BytesIO(b"not excel"), "bad.txt")},
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(response.get_json()["success"])

    def test_price_audit_upload_and_start_validation_are_available(self):
        base_dir = Path(tempfile.mkdtemp())
        app = create_app(base_dir=base_dir)
        client = app.test_client()

        upload_response = client.post(
            "/api/upload",
            data={"file": (io.BytesIO(self._price_workbook_bytes()), "测价.xlsx")},
            content_type="multipart/form-data",
        )

        self.assertEqual(upload_response.status_code, 200)
        upload_payload = upload_response.get_json()
        self.assertTrue(upload_payload["success"])
        self.assertTrue(Path(upload_payload["path"]).exists())

        invalid_threshold_response = client.post(
            "/api/start",
            json={"file": upload_payload["path"], "threshold": "abc"},
        )
        self.assertEqual(invalid_threshold_response.status_code, 400)
        self.assertIn("价格门槛", invalid_threshold_response.get_json()["error"])

        status_response = client.get("/api/status")
        self.assertEqual(status_response.status_code, 200)
        self.assertIn("running", status_response.get_json())

    def test_price_audit_captures_low_price_screenshots_before_writing_results(self):
        source = Path("app.py").read_text(encoding="utf-8")

        self.assertIn("capture_low_price_result_screenshots_with_page_factory", source)
        self.assertLess(
            source.index("capture_low_price_result_screenshots_with_page_factory("),
            source.index("write_results("),
        )

    def test_price_audit_worker_browser_visibility_is_configurable(self):
        source = Path("app.py").read_text(encoding="utf-8")

        worker_factory_source = source[source.index("def create_worker_page"):source.index("def on_result")]
        self.assertIn("show_browser = bool(data.get(\"show_browser\"))", source)
        self.assertIn("args=(input_file, threshold, show_browser, concurrent_workers)", source)
        self.assertIn("worker_headless = not show_browser", source)
        self.assertIn("headless=worker_headless", worker_factory_source)
        self.assertIn("concurrent_workers", source)

    def test_price_audit_scan_workers_block_images_but_screenshot_workers_keep_images(self):
        source = Path("app.py").read_text(encoding="utf-8")

        self.assertIn("block_images=block_images", source)
        self.assertLess(
            source.index("block_images=True"),
            source.rindex("block_images=False"),
        )

    def test_price_audit_closes_login_browser_and_returns_to_worker_browser_mode_after_login(self):
        source = Path("app.py").read_text(encoding="utf-8")

        self.assertGreaterEqual(
            source.count("headless=worker_headless"),
            2,
        )
        self.assertIn("登录成功后切回测价浏览器", source)
        self.assertIn("检查测价浏览器登录状态", source)
        self.assertIn("登录状态未能同步到测价浏览器", source)
        self.assertIn("低价截图：应补", source)

    def test_price_audit_login_browser_disables_resource_blocking(self):
        source = Path("app.py").read_text(encoding="utf-8")

        self.assertGreaterEqual(
            source.count("block_resources=False"),
            2,
        )

    def test_stop_request_closes_active_browsers(self):
        source = Path("app.py").read_text(encoding="utf-8")

        self.assertIn("def close_current_browsers", source)
        self.assertIn("close_current_browsers()", source[source.index("def stop_price_audit"):])


    def test_bigscreen_capture_preview_returns_room_id_and_hours(self):
        app = create_app(base_dir=Path(tempfile.mkdtemp()))
        client = app.test_client()

        response = client.post(
            "/api/bigscreen-capture/preview",
            json={"url": "https://jlive.jd.com/bigScreen?id=46794566"},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["room_id"], "46794566")
        self.assertEqual(payload["hour_options"][0], "10:00")
        self.assertIn("10:30", payload["hour_options"])
        self.assertIn("19:00", payload["hour_options"])
        self.assertIn("19:30", payload["hour_options"])
        self.assertIn("24:00", payload["hour_options"])
        self.assertEqual(payload["hour_options"][-1], "24:00")

    def test_bigscreen_capture_rejects_bad_url(self):
        app = create_app(base_dir=Path(tempfile.mkdtemp()))
        client = app.test_client()

        response = client.post(
            "/api/bigscreen-capture/preview",
            json={"url": "https://example.com"},
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(response.get_json()["success"])

    def test_bigscreen_capture_capture_now_starts_async_task_and_downloads_zip(self):
        class FakeCaptureResult:
            room_id = "46794566"
            success_count = 15
            fail_count = 0
            stopped_count = 0
            records = [object()]

        class ImmediateThread:
            def __init__(self, target, args):
                self.target = target
                self.args = args

            def start(self):
                self.target(*self.args)

        app = create_app(base_dir=Path(tempfile.mkdtemp()))
        client = app.test_client()

        with patch("app.capture_bigscreen_once") as fake_capture, \
                patch("app.write_bigscreen_capture_bundle") as fake_bundle, \
                patch("app.threading.Thread", ImmediateThread):
            zip_file = app.config["BIGSCREEN_OUTPUT_DIR"] / "result.zip"
            zip_file.write_bytes(b"zip")
            FakeCaptureResult.zip_file = zip_file
            fake_capture.return_value = FakeCaptureResult
            fake_bundle.return_value = (app.config["BIGSCREEN_OUTPUT_DIR"] / "manifest.xlsx", zip_file)

            response = client.post(
                "/api/bigscreen-capture/capture-now",
                json={"url": "https://jlive.jd.com/bigScreen?id=46794566", "show_browser": True},
            )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(fake_capture.call_args.kwargs["show_browser"])
        self.assertEqual(fake_capture.call_args.kwargs["planned_slot"], "立即截图")
        payload = response.get_json()
        self.assertTrue(payload["success"])
        self.assertIn("task_id", payload)
        status_payload = client.get("/api/bigscreen-capture/status").get_json()
        self.assertFalse(status_payload["running"])
        self.assertEqual(status_payload["success_count"], 15)
        download_response = client.get(f"/api/bigscreen-capture/download/{payload['task_id']}")
        self.assertEqual(download_response.status_code, 200)
        self.assertEqual(download_response.data, b"zip")
        download_response.close()

    def test_bigscreen_capture_download_recovers_zip_from_task_directory(self):
        app = create_app(base_dir=Path(tempfile.mkdtemp()))
        client = app.test_client()
        task_dir = app.config["BIGSCREEN_OUTPUT_DIR"] / "lost-task"
        slot_dir = task_dir / "1800"
        slot_dir.mkdir(parents=True)
        (task_dir / ".DS_Store").write_bytes(b"mac")
        (slot_dir / "截图清单.xlsx").write_bytes(b"manifest")
        (slot_dir / "sample.png").write_bytes(b"png")

        response = client.get("/api/bigscreen-capture/download/lost-task")

        self.assertEqual(response.status_code, 200)
        with ZipFile(io.BytesIO(response.data)) as archive:
            names = archive.namelist()
        self.assertIn("1800/截图清单.xlsx", names)
        self.assertIn("1800/sample.png", names)
        self.assertNotIn(".DS_Store", names)
        response.close()

    def test_bigscreen_capture_capture_now_rejects_duplicate_running_task(self):
        class FakeCaptureResult:
            room_id = "46794566"
            success_count = 15
            fail_count = 0
            stopped_count = 0
            records = [object()]

        class IdleThread:
            def __init__(self, target, args):
                pass

            def start(self):
                pass

        app = create_app(base_dir=Path(tempfile.mkdtemp()))
        client = app.test_client()

        with patch("app.capture_bigscreen_once") as fake_capture, \
                patch("app.threading.Thread", IdleThread):
            zip_file = app.config["BIGSCREEN_OUTPUT_DIR"] / "result.zip"
            zip_file.write_bytes(b"zip")
            FakeCaptureResult.zip_file = zip_file
            fake_capture.return_value = FakeCaptureResult
            first_response = client.post(
                "/api/bigscreen-capture/capture-now",
                json={"url": "https://jlive.jd.com/bigScreen?id=46794566"},
            )
            second_response = client.post(
                "/api/bigscreen-capture/capture-now",
                json={"url": "https://jlive.jd.com/bigScreen?id=46794566"},
            )

        self.assertEqual(first_response.status_code, 200)
        self.assertEqual(second_response.status_code, 400)
        self.assertIn("已有蓝屏截图任务正在运行", second_response.get_json()["error"])

    def test_bigscreen_capture_start_rejects_empty_slots(self):
        app = create_app(base_dir=Path(tempfile.mkdtemp()))
        client = app.test_client()

        response = client.post(
            "/api/bigscreen-capture/start",
            json={"url": "https://jlive.jd.com/bigScreen?id=46794566", "hour_slots": []},
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(response.get_json()["success"])
        self.assertEqual(response.get_json()["error"], "请选择至少一个时间点")

    def test_bigscreen_capture_start_uses_time_point_error_messages(self):
        app = create_app(base_dir=Path(tempfile.mkdtemp()))
        client = app.test_client()

        invalid_response = client.post(
            "/api/bigscreen-capture/start",
            json={
                "url": "https://jlive.jd.com/bigScreen?id=46794566",
                "hour_slots": ["bad"],
            },
        )
        expired_response = client.post(
            "/api/bigscreen-capture/start",
            json={
                "url": "https://jlive.jd.com/bigScreen?id=46794566",
                "capture_date": (date.today() - timedelta(days=1)).isoformat(),
                "hour_slots": ["10:00"],
            },
        )

        self.assertEqual(invalid_response.status_code, 400)
        self.assertEqual(invalid_response.get_json()["error"], "时间点配置无效")
        self.assertEqual(expired_response.status_code, 400)
        self.assertEqual(expired_response.get_json()["error"], "选择的时间点都已过期")

    def test_bigscreen_capture_start_cleans_runtime_before_scheduling(self):
        class IdleThread:
            def __init__(self, target, args):
                pass

            def start(self):
                pass

        app = create_app(base_dir=Path(tempfile.mkdtemp()))
        client = app.test_client()

        with patch("app._cleanup_runtime_for_app") as cleanup, \
                patch("app.threading.Thread", IdleThread):
            response = client.post(
                "/api/bigscreen-capture/start",
                json={
                    "url": "https://jlive.jd.com/bigScreen?id=46794566",
                    "capture_date": (date.today() + timedelta(days=1)).isoformat(),
                    "hour_slots": ["10:00"],
                },
            )

        self.assertEqual(response.status_code, 200)
        cleanup.assert_called_once_with(app)

    def test_bigscreen_capture_status_returns_initial_shape(self):
        app = create_app(base_dir=Path(tempfile.mkdtemp()))
        client = app.test_client()

        response = client.get("/api/bigscreen-capture/status")

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertFalse(payload["running"])
        self.assertEqual(payload["total"], 0)
        self.assertEqual(payload["logs"], [])

    def test_bigscreen_capture_stop_sets_stopping(self):
        app = create_app(base_dir=Path(tempfile.mkdtemp()))
        client = app.test_client()

        stop_response = client.post("/api/bigscreen-capture/stop")
        status_response = client.get("/api/bigscreen-capture/status")

        self.assertEqual(stop_response.status_code, 200)
        self.assertTrue(status_response.get_json()["stopping"])

    def test_bigscreen_capture_source_uses_combined_bundle_login_and_show_browser(self):
        source = Path("app.py").read_text(encoding="utf-8")

        self.assertIn("show_browser = bool(data.get(\"show_browser\", False))", source)
        self.assertIn("show_browser=show_browser", source)
        self.assertIn("write_bigscreen_capture_bundle(", source)
        self.assertIn("bigscreen_status[\"need_login\"] = True", source)
        self.assertIn("bigscreen_login_event.wait", source)
        self.assertIn("stopped_count", source)

    def test_room_creator_preview_stores_column_mapping(self):
        app = create_app(base_dir=Path(tempfile.mkdtemp()))
        client = app.test_client()

        wb = Workbook()
        ws = wb.active
        ws.append(["直播标题", "开播时间", "直播形式", "画面方向", "直播地点", "直播品类"])
        ws.append(["直播间一号", "2026-07-01 20:00:00", "测试直播", "横屏", "不显示地点", "母婴"])
        output = io.BytesIO()
        wb.save(output)
        wb.close()

        response = client.post(
            "/api/room-creator/preview",
            data={"file": (io.BytesIO(output.getvalue()), "rooms.xlsx")},
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["mapping"]["title_col"], "直播标题")
        self.assertEqual(payload["mapping"]["start_time_col"], "开播时间")
        self.assertEqual(payload["mapping"]["live_form_col"], "直播形式")
        self.assertEqual(payload["mapping"]["live_direction_col"], "画面方向")
        self.assertEqual(payload["mapping"]["live_location_col"], "直播地点")
        self.assertEqual(payload["mapping"]["live_category_col"], "直播品类")
        self.assertIn(payload["task_id"], app.config["ROOM_CREATOR_MAPPINGS"])
    def test_parse_sku_input_supports_multiple_separators_and_dedup(self):
        self.assertEqual(
            parse_sku_input("100264886683,48279162646;100264886683"),
            ["100264886683", "48279162646"],
        )
        self.assertEqual(
            parse_sku_input("100264886683，48279162646；100264886683"),
            ["100264886683", "48279162646"],
        )
        self.assertEqual(
            parse_sku_input("100264886683\n48279162646\t100264886683"),
            ["100264886683", "48279162646"],
        )

    def test_start_from_skus_rejects_invalid_threshold_with_json_error(self):
        base_dir = Path(tempfile.mkdtemp())
        app = create_app(base_dir=base_dir)
        client = app.test_client()

        response = client.post(
            "/api/start-from-skus",
            json={"skus": "100264886683", "threshold": "abc"},
        )
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.get_json()["success"], False)
        self.assertIn("价格门槛", response.get_json()["error"])

    def test_start_from_skus_rejects_empty_input(self):
        base_dir = Path(tempfile.mkdtemp())
        app = create_app(base_dir=base_dir)
        client = app.test_client()

        response = client.post(
            "/api/start-from-skus",
            json={"skus": "", "threshold": 6},
        )
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.get_json()["success"], False)
        self.assertIn("SKU", response.get_json()["error"])

    @patch("threading.Thread")
    def test_start_from_skus_creates_input_file_and_returns_count(self, mock_thread):
        base_dir = Path(tempfile.mkdtemp())
        app = create_app(base_dir=base_dir)
        client = app.test_client()

        response = client.post(
            "/api/start-from-skus",
            json={"skus": "100264886683,48279162646", "threshold": 6, "concurrent_workers": 1},
        )
        self.assertEqual(response.status_code, 200)
        data = response.get_json()
        self.assertTrue(data["success"])
        self.assertEqual(data["count"], 2)

        input_dir = app.config["PRICE_INPUT_DIR"]
        generated_files = list(input_dir.glob("页面输入SKU_*.xlsx"))
        self.assertTrue(generated_files, "应生成临时输入文件")
        for f in generated_files:
            f.unlink(missing_ok=True)

        mock_thread.assert_called_once()
        self.assertEqual(mock_thread.call_args.kwargs.get("kwargs"), {"cleanup_input": True})

    def test_start_from_skus_requests_cleanup_of_temp_input_file(self):
        source = Path("app.py").read_text(encoding="utf-8")
        self.assertIn("cleanup_input", source)
        self.assertIn('kwargs={"cleanup_input": True}', source)
        self.assertIn("_unlink_with_retries(input_file)", source)

    def test_unlink_with_retries_handles_transient_windows_file_lock(self):
        path = Path("locked.xlsx")

        with patch.object(Path, "unlink", side_effect=[PermissionError("busy"), None]) as unlink, \
                patch("app.time.sleep") as sleep:
            self.assertTrue(_unlink_with_retries(path, attempts=2, delay_seconds=0.01))

        self.assertEqual(unlink.call_count, 2)
        sleep.assert_called_once_with(0.01)

    def _business_workbook_bytes(self):
        wb = Workbook()
        ws = wb.active
        ws.append(
            [
                "skuID",
                "商品名称",
                "促销编码 专享券填专享券key码 专享价填ERP促销编码 （达人id：22766602）",
            ]
        )
        ws.append(["1001", "A 商品", "vender_BA#a9d94c41368e441094132b17a3b40fd6"])
        ws.append(["1002", "B 商品", "381421541016"])
        output = io.BytesIO()
        wb.save(output)
        wb.close()
        return output.getvalue()

    def _manual_mapping_workbook_bytes(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["排期", "商品编号", "标题", "权益内容"])
        ws.append(["2026-06-18", "1001", "A 商品", "vender_BA#a9d94c41368e441094132b17a3b40fd6"])
        ws.append(["2026-06-19", "1002", "B 商品", "381421541016"])
        output = io.BytesIO()
        wb.save(output)
        wb.close()
        return output.getvalue()

    def _business_workbook_with_selling_point_bytes(self):
        wb = Workbook()
        ws = wb.active
        ws.append(
            [
                "skuID",
                "商品名称",
                "促销编码 专享券填专享券key码 专享价填ERP促销编码",
                "短卖点（折扣、直降、卖点都可以）",
            ]
        )
        ws.append(["1001", "A 商品", "vender_BA#a9d94c41368e441094132b17a3b40fd6", "限时直降"])
        ws.append(["1002", "B 商品", "381421541016", "满减优惠"])
        output = io.BytesIO()
        wb.save(output)
        wb.close()
        return output.getvalue()

    def _price_workbook_bytes(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["商品SKU"])
        ws.append(["100264886683"])
        output = io.BytesIO()
        wb.save(output)
        wb.close()
        return output.getvalue()


if __name__ == "__main__":
    unittest.main()
