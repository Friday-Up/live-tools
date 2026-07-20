import io
import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

from app import create_app


class RedRainRoutesTest(unittest.TestCase):
    def workbook_bytes(self):
        output = io.BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["活动名称", "开始时间", "结束时间", "红包发放方式", "红包ID", "中奖概率"])
        sheet.append(["晚场红包雨", datetime(2099, 7, 20, 20), datetime(2099, 7, 20, 20, 10), "普通发放", "123456", 50])
        workbook.save(output)
        workbook.close()
        return output.getvalue()

    def mixed_workbook_bytes(self):
        output = io.BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["活动名称", "开始时间", "结束时间", "红包发放方式", "红包ID", "中奖概率"])
        sheet.append(["坏行", "not-a-time", datetime(2099, 7, 20, 20, 10), "普通发放", "bad", 50])
        sheet.append(["正常行", datetime(2099, 7, 20, 21), datetime(2099, 7, 20, 21, 10), "普通发放", "good", 50])
        workbook.save(output)
        workbook.close()
        return output.getvalue()

    def empty_workbook_bytes(self):
        output = io.BytesIO()
        workbook = Workbook()
        workbook.active.append(["活动名称", "开始时间", "结束时间", "红包发放方式", "红包ID", "中奖概率"])
        workbook.save(output)
        workbook.close()
        return output.getvalue()

    def test_preview_and_template_routes(self):
        app = create_app(base_dir=Path(tempfile.mkdtemp()))
        client = app.test_client()
        response = client.post(
            "/api/red-rain/preview",
            data={"file": (io.BytesIO(self.workbook_bytes()), "红包雨.xlsx")},
            content_type="multipart/form-data",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["row_count"], 1)
        self.assertEqual(payload["invalid_count"], 0)
        self.assertEqual(payload["mapping"]["red_packet_id_col"], "红包ID")

        template = client.get("/api/red-rain/template")
        self.assertEqual(template.status_code, 200)
        self.assertGreater(len(template.data), 1000)
        template.close()

    def test_start_reserves_task_before_worker_thread_runs(self):
        app = create_app(base_dir=Path(tempfile.mkdtemp()))
        client = app.test_client()
        preview = client.post(
            "/api/red-rain/preview",
            data={"file": (io.BytesIO(self.workbook_bytes()), "红包雨.xlsx")},
            content_type="multipart/form-data",
        ).get_json()

        with patch("app.threading.Thread") as thread:
            first = client.post("/api/red-rain/start", json={"task_id": preview["task_id"]})
            second = client.post("/api/red-rain/start", json={"task_id": preview["task_id"]})

        self.assertEqual(first.status_code, 200)
        self.assertTrue(first.get_json()["success"])
        self.assertEqual(second.status_code, 400)
        self.assertIn("已有红包雨创建任务", second.get_json()["error"])
        thread.return_value.start.assert_called_once_with()

    def test_preview_counts_bad_rows_without_rejecting_whole_file(self):
        app = create_app(base_dir=Path(tempfile.mkdtemp()))
        response = app.test_client().post(
            "/api/red-rain/preview",
            data={"file": (io.BytesIO(self.mixed_workbook_bytes()), "混合数据.xlsx")},
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["row_count"], 2)
        self.assertEqual(payload["invalid_count"], 1)

    def test_preview_rejects_empty_workbook(self):
        app = create_app(base_dir=Path(tempfile.mkdtemp()))
        response = app.test_client().post(
            "/api/red-rain/preview",
            data={"file": (io.BytesIO(self.empty_workbook_bytes()), "空表.xlsx")},
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("没有可执行数据", response.get_json()["error"])

    def test_room_creator_template_route(self):
        app = create_app(base_dir=Path(tempfile.mkdtemp()))
        response = app.test_client().get("/api/room-creator/template")

        self.assertEqual(response.status_code, 200)
        self.assertGreater(len(response.data), 1000)
        response.close()


if __name__ == "__main__":
    unittest.main()
