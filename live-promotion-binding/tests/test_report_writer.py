import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from promotion_binding.report_writer import IssueRecord, write_report
from promotion_binding.template_writer import BindingRecord, BindingType


class ReportWriterTest(unittest.TestCase):
    def test_writes_business_friendly_report_sheets(self):
        output_path = Path(tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name)

        write_report(
            output_path=output_path,
            summary={
                "可绑定条数": 1,
                "异常条数": 1,
                "空值跳过": 2,
            },
            binding_records=[
                BindingRecord(
                    sku="100089021178",
                    binding_type=BindingType.PROMO_ID,
                    binding_value="381421541016",
                    source_row=8,
                    product_name="A 商品",
                )
            ],
            issue_records=[
                IssueRecord(
                    source_row=25,
                    sku="100223622026",
                    product_name="B 商品",
                    raw_code="百亿补贴",
                    issue_type="INVALID_CODE",
                    message="未识别到专享券KEY或专享价促销ID",
                    action="人工确认后重填",
                ),
                IssueRecord(
                    source_row=26,
                    sku="100089021178",
                    product_name="C 商品",
                    raw_code="381421541016",
                    issue_type="DUPLICATE_BINDING",
                    message="重复SKU和绑定值，已只保留第一条",
                    action="无需处理；如需调整顺序请修改原表",
                    kept_source_row=8,
                ),
            ],
        )

        wb = load_workbook(output_path, data_only=True)

        self.assertEqual(wb.sheetnames, ["汇总", "可上传明细", "需处理异常", "跳过和重复"])
        self.assertEqual([wb["汇总"].cell(1, col).value for col in range(1, 5)], ["分类", "指标", "数量", "说明"])
        self.assertEqual(wb["汇总"].cell(2, 2).value, "可绑定条数")
        self.assertEqual(wb["汇总"].cell(2, 3).value, 1)
        self.assertEqual(
            [wb["可上传明细"].cell(1, col).value for col in range(1, 7)],
            ["原始行号", "SKU", "商品名称", "绑定类型", "绑定值", "写入模板行号"],
        )
        self.assertEqual(wb["可上传明细"].cell(2, 1).value, 8)
        self.assertEqual(wb["可上传明细"].cell(2, 3).value, "A 商品")
        self.assertEqual(wb["可上传明细"].cell(2, 4).value, "专享价促销ID")
        self.assertEqual(wb["可上传明细"].cell(2, 6).value, 2)
        self.assertEqual(wb["需处理异常"].cell(2, 1).value, 25)
        self.assertEqual(wb["需处理异常"].cell(2, 3).value, "B 商品")
        self.assertEqual(wb["需处理异常"].cell(2, 5).value, "未识别到有效券码/促销ID")
        self.assertEqual(wb["跳过和重复"].cell(2, 1).value, 26)
        self.assertEqual(wb["跳过和重复"].cell(2, 5).value, "重复 SKU 和绑定值")
        self.assertEqual(wb["跳过和重复"].cell(2, 6).value, 8)


if __name__ == "__main__":
    unittest.main()
