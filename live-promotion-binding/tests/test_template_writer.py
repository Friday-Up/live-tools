import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from promotion_binding.template_writer import BindingRecord, BindingType, write_upload_template


ASSET_TEMPLATE = (
    Path(__file__).resolve().parents[1]
    / "assets"
    / "商品上传模版（2026切片版）.xlsx"
)


class TemplateWriterTest(unittest.TestCase):
    def test_writes_binding_records_to_official_template_columns(self):
        output_path = Path(tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name)

        write_upload_template(
            template_path=ASSET_TEMPLATE,
            output_path=output_path,
            records=[
                BindingRecord(
                    sku="10079660739051",
                    binding_type=BindingType.COUPON_KEY,
                    binding_value="vender_BA#a9d94c41368e441094132b17a3b40fd6",
                    source_row=3,
                ),
                BindingRecord(
                    sku="100089021178",
                    binding_type=BindingType.PROMO_ID,
                    binding_value="381421541016",
                    source_row=8,
                ),
            ],
        )

        wb = load_workbook(output_path, data_only=False)
        ws = wb.active

        self.assertEqual(ws.cell(1, 1).value, "SKUID（必填，填写格式请用文本/数字格式，勿使用科学计数格式）")
        self.assertEqual(ws.cell(2, 1).value, "10079660739051")
        self.assertEqual(ws.cell(2, 3).value, "vender_BA#a9d94c41368e441094132b17a3b40fd6")
        self.assertIsNone(ws.cell(2, 4).value)
        self.assertEqual(ws.cell(3, 1).value, "100089021178")
        self.assertIsNone(ws.cell(3, 3).value)
        self.assertEqual(ws.cell(3, 4).value, "381421541016")
        self.assertEqual(ws.cell(2, 1).number_format, "@")
        self.assertEqual(ws.cell(3, 4).number_format, "@")

    def test_writes_selling_point_to_template_b_column(self):
        output_path = Path(tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name)

        write_upload_template(
            template_path=ASSET_TEMPLATE,
            output_path=output_path,
            records=[
                BindingRecord(
                    sku="1001",
                    binding_type=BindingType.COUPON_KEY,
                    binding_value="vender_BA#a9d94c41368e441094132b17a3b40fd6",
                    source_row=2,
                    selling_point="限时直降",
                ),
                BindingRecord(
                    sku="1002",
                    binding_type=None,
                    binding_value="",
                    source_row=3,
                    selling_point="满减优惠",
                ),
            ],
        )

        wb = load_workbook(output_path, data_only=False)
        ws = wb.active
        self.assertEqual(ws.cell(2, 1).value, "1001")
        self.assertEqual(ws.cell(2, 2).value, "限时直降")
        self.assertEqual(ws.cell(2, 3).value, "vender_BA#a9d94c41368e441094132b17a3b40fd6")
        self.assertEqual(ws.cell(3, 1).value, "1002")
        self.assertEqual(ws.cell(3, 2).value, "满减优惠")
        self.assertIsNone(ws.cell(3, 3).value)


if __name__ == "__main__":
    unittest.main()
