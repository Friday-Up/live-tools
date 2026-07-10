import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from promotion_binding.service import generate_binding_files


ASSET_TEMPLATE = (
    Path(__file__).resolve().parents[1]
    / "assets"
    / "商品上传模版（2026切片版）.xlsx"
)


class ServiceTest(unittest.TestCase):
    def test_generates_template_and_report_with_dedupe_and_issues(self):
        business_path = self._create_business_workbook()
        output_dir = Path(tempfile.mkdtemp())

        result = generate_binding_files(
            business_file=business_path,
            template_file=ASSET_TEMPLATE,
            output_dir=output_dir,
            generated_at=datetime(2026, 6, 18, 14, 20, 26),
        )

        self.assertEqual(result.success_count, 2)
        self.assertEqual(result.coupon_key_count, 1)
        self.assertEqual(result.promo_id_count, 1)
        self.assertEqual(result.skipped_empty_count, 1)
        self.assertEqual(result.invalid_count, 3)
        self.assertEqual(result.duplicate_count, 1)
        self.assertTrue(result.output_template_path.exists())
        self.assertTrue(result.report_path.exists())
        self.assertEqual(result.output_template_path.name, "京东绑券上传模板_20260618-142026_.xlsx")
        self.assertEqual(result.report_path.name, "异常报告_20260618-142026_.xlsx")

        template_wb = load_workbook(result.output_template_path, data_only=True)
        try:
            template_ws = template_wb.active
            self.assertEqual(template_ws.cell(2, 1).value, "1001")
            self.assertEqual(template_ws.cell(2, 3).value, "vender_BA#a9d94c41368e441094132b17a3b40fd6")
            self.assertEqual(template_ws.cell(3, 1).value, "1002")
            self.assertEqual(template_ws.cell(3, 4).value, "381421541016")
            self.assertIsNone(template_ws.cell(4, 1).value)
        finally:
            template_wb.close()

        report_wb = load_workbook(result.report_path, data_only=True)
        try:
            self.assertEqual(report_wb.sheetnames, ["汇总", "可上传明细", "需处理异常", "跳过和重复", "短卖点警告"])
            self.assertEqual(report_wb["可上传明细"].cell(2, 3).value, "商品 1001")
            self.assertEqual(report_wb["可上传明细"].cell(3, 7).value, 3)

            manual_issue_rows = [
                [
                    report_wb["需处理异常"].cell(row, col).value
                    for col in range(1, report_wb["需处理异常"].max_column + 1)
                ]
                for row in range(2, report_wb["需处理异常"].max_row + 1)
            ]
            manual_issue_messages = [row[4] for row in manual_issue_rows]
            self.assertIn("未识别到有效券码/促销ID", manual_issue_messages)
            self.assertIn("同一 SKU 出现多个不同绑定值", manual_issue_messages)
            self.assertIn("同一单元格有多个券码 KEY", manual_issue_messages)

            skipped_rows = [
                [
                    report_wb["跳过和重复"].cell(row, col).value
                    for col in range(1, report_wb["跳过和重复"].max_column + 1)
                ]
                for row in range(2, report_wb["跳过和重复"].max_row + 1)
            ]
            self.assertIn("未填写券码/价码", [row[4] for row in skipped_rows])
            duplicate_row = next(row for row in skipped_rows if row[1] == "1002")
            self.assertEqual(duplicate_row[2], "商品 1002 重复")
            self.assertEqual(duplicate_row[4], "重复 SKU 和绑定值")
            self.assertEqual(duplicate_row[5], 3)
        finally:
            report_wb.close()

    def _create_business_workbook(self):
        path = Path(tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name)
        wb = Workbook()
        ws = wb.active
        ws.append(["商品名称*【必填】", "上播·SKU ID*【必填】", "券码/价码 达人id：22766602"])
        ws.append(["商品 1001", "1001", "vender_BA#a9d94c41368e441094132b17a3b40fd6"])
        ws.append(["商品 1002", "1002", "381421541016"])
        ws.append(["商品 1003", "1003", "百补"])
        ws.append(["商品 1004", "1004", None])
        ws.append(["商品 1002 重复", "1002", "381421541016"])
        ws.append(["商品 1005 A", "1005", "381421541016"])
        ws.append(["商品 1005 B", "1005", "402673865206"])
        ws.append(
            [
                "商品 1006",
                "1006",
                "vender_BA#a825d4ab8f8f4ba6960aa20c142dabbb；"
                "vender_BA#65482cc4e1a24ce99acbe89014f3530f",
            ]
        )
        wb.save(path)
        wb.close()
        return path

    def test_includes_selling_point_when_switch_enabled(self):
        path = Path(tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name)
        wb = Workbook()
        ws = wb.active
        ws.append(["skuID", "商品名称", "券码", "短卖点（折扣、直降、卖点都可以）"])
        ws.append(["1001", "A 商品", "vender_BA#a9d94c41368e441094132b17a3b40fd6", "限时直降"])
        ws.append(["1002", "B 商品", "381421541016", "满减优惠"])
        wb.save(path)
        wb.close()

        output_dir = Path(tempfile.mkdtemp())
        result = generate_binding_files(
            business_file=path,
            template_file=ASSET_TEMPLATE,
            output_dir=output_dir,
            enable_selling_point=True,
        )

        self.assertTrue(result.selling_point_column_found)
        self.assertEqual(result.success_count, 2)
        self.assertEqual(result.selling_point_count, 2)

        template_wb = load_workbook(result.output_template_path, data_only=True)
        try:
            template_ws = template_wb.active
            self.assertEqual(template_ws.cell(2, 2).value, "限时直降")
            self.assertEqual(template_ws.cell(3, 2).value, "满减优惠")
        finally:
            template_wb.close()

        report_wb = load_workbook(result.report_path, data_only=True)
        try:
            self.assertIn("短卖点警告", report_wb.sheetnames)
            summary_values = {report_wb["汇总"].cell(row, 2).value: report_wb["汇总"].cell(row, 3).value for row in range(2, report_wb["汇总"].max_row + 1)}
            self.assertEqual(summary_values["含短卖点条数"], 2)
        finally:
            report_wb.close()

    def test_includes_selling_point_only_rows(self):
        path = Path(tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name)
        wb = Workbook()
        ws = wb.active
        ws.append(["skuID", "商品名称", "券码", "短卖点"])
        ws.append(["1001", "A 商品", None, "仅短卖点"])
        ws.append(["1002", "B 商品", "381421541016", "有价码和短卖点"])
        wb.save(path)
        wb.close()

        output_dir = Path(tempfile.mkdtemp())
        result = generate_binding_files(
            business_file=path,
            template_file=ASSET_TEMPLATE,
            output_dir=output_dir,
            enable_selling_point=True,
        )

        self.assertEqual(result.success_count, 2)
        self.assertEqual(result.promo_id_count, 1)
        self.assertEqual(result.coupon_key_count, 0)
        self.assertEqual(result.selling_point_count, 2)

        template_wb = load_workbook(result.output_template_path, data_only=True)
        try:
            template_ws = template_wb.active
            self.assertEqual(template_ws.cell(2, 1).value, "1001")
            self.assertEqual(template_ws.cell(2, 2).value, "仅短卖点")
            self.assertIsNone(template_ws.cell(2, 4).value)
            self.assertEqual(template_ws.cell(3, 1).value, "1002")
            self.assertEqual(template_ws.cell(3, 2).value, "有价码和短卖点")
            self.assertEqual(template_ws.cell(3, 4).value, "381421541016")
        finally:
            template_wb.close()

    def test_reports_selling_point_too_long_as_warning(self):
        path = Path(tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name)
        wb = Workbook()
        ws = wb.active
        ws.append(["skuID", "商品名称", "券码", "短卖点"])
        ws.append(["1001", "A 商品", "381421541016", "这是一个超过二十二个字符的短卖点内容示例很好啊"])
        wb.save(path)
        wb.close()

        output_dir = Path(tempfile.mkdtemp())
        result = generate_binding_files(
            business_file=path,
            template_file=ASSET_TEMPLATE,
            output_dir=output_dir,
            enable_selling_point=True,
        )

        self.assertEqual(result.success_count, 1)
        self.assertEqual(result.selling_point_count, 1)
        self.assertEqual(result.invalid_count, 0)

        report_wb = load_workbook(result.report_path, data_only=True)
        try:
            warning_ws = report_wb["短卖点警告"]
            self.assertEqual(warning_ws.max_row, 2)
            self.assertEqual(warning_ws.cell(2, 4).value, "这是一个超过二十二个字符的短卖点内容示例很好啊")
        finally:
            report_wb.close()

        template_wb = load_workbook(result.output_template_path, data_only=True)
        try:
            self.assertEqual(template_wb.active.cell(2, 2).value, "这是一个超过二十二个字符的短卖点内容示例很好啊")
        finally:
            template_wb.close()


if __name__ == "__main__":
    unittest.main()
