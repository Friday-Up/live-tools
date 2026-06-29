from __future__ import annotations

"""Write JD live upload templates for promotion bindings."""

from dataclasses import dataclass
from enum import Enum
from pathlib import Path

from openpyxl import load_workbook


class BindingType(str, Enum):
    COUPON_KEY = "COUPON_KEY"
    PROMO_ID = "PROMO_ID"


@dataclass(frozen=True)
class BindingRecord:
    sku: str
    binding_type: BindingType
    binding_value: str
    source_row: int
    product_name: str = ""
    selling_point: str = ""


def write_upload_template(
    template_path: str | Path,
    output_path: str | Path,
    records: list[BindingRecord],
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(template_path)
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        for col in range(1, 8):
            ws.cell(row, col).value = None

    for index, record in enumerate(records, start=2):
        sku_cell = ws.cell(index, 1)
        sku_cell.value = str(record.sku)
        sku_cell.number_format = "@"

        if record.selling_point:
            selling_point_cell = ws.cell(index, 2)
            selling_point_cell.value = str(record.selling_point)

        if record.binding_type == BindingType.COUPON_KEY:
            key_cell = ws.cell(index, 3)
            key_cell.value = str(record.binding_value)
            key_cell.number_format = "@"
        elif record.binding_type == BindingType.PROMO_ID:
            promo_cell = ws.cell(index, 4)
            promo_cell.value = str(record.binding_value)
            promo_cell.number_format = "@"

    wb.save(output_path)
    return output_path
