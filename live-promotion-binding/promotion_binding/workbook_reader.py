from __future__ import annotations

"""Read business submission workbooks for promotion binding generation."""

from dataclasses import dataclass
from pathlib import Path
import re

from openpyxl import load_workbook


@dataclass(frozen=True)
class BusinessRow:
    source_row: int
    sku: str
    raw_code: str
    product_name: str = ""
    selling_point: str = ""


@dataclass(frozen=True)
class ColumnMapping:
    sku_col: int | None = None
    code_col: int | None = None
    product_name_col: int | None = None
    selling_point_col: int | None = None


@dataclass(frozen=True)
class WorkbookColumn:
    index: int
    header: str
    sample_values: list[str]


@dataclass(frozen=True)
class WorkbookInspection:
    columns: list[WorkbookColumn]
    suggested_mapping: ColumnMapping


SKU_COLUMN_KEYWORDS = ["sku", "skuid", "上播skuid"]
CODE_COLUMN_KEYWORDS = ["券码", "价码", "促销编码", "专享券", "专享价", "达人id"]
PRODUCT_NAME_COLUMN_KEYWORDS = ["商品名称"]
SELLING_POINT_COLUMN_KEYWORDS = ["短卖点", "利益点", "卖点"]


def normalize_header(value) -> str:
    text = "" if value is None else str(value)
    return re.sub(r"[\s\u00a0]+", "", text).strip().casefold()


def format_cell_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def _find_column(headers: list, candidates: list[str]) -> int | None:
    normalized_headers = [normalize_header(header) for header in headers]
    for index, header in enumerate(normalized_headers):
        for candidate in candidates:
            if normalize_header(candidate) in header:
                return index + 1
    return None


def inspect_business_workbook(file_path: str | Path, sample_size: int = 3) -> WorkbookInspection:
    wb = load_workbook(file_path, data_only=True, read_only=True)
    try:
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        columns: list[WorkbookColumn] = []
        for col_index, header in enumerate(headers, start=1):
            sample_values: list[str] = []
            for row_index in range(2, min(ws.max_row, sample_size + 1) + 1):
                value = format_cell_value(ws.cell(row_index, col_index).value)
                if value:
                    sample_values.append(value)
            columns.append(
                WorkbookColumn(
                    index=col_index,
                    header=format_cell_value(header),
                    sample_values=sample_values,
                )
            )

        return WorkbookInspection(
            columns=columns,
            suggested_mapping=ColumnMapping(
                sku_col=_find_column(headers, SKU_COLUMN_KEYWORDS),
                code_col=_find_column(headers, CODE_COLUMN_KEYWORDS),
                product_name_col=_find_column(headers, PRODUCT_NAME_COLUMN_KEYWORDS),
                selling_point_col=_find_column(headers, SELLING_POINT_COLUMN_KEYWORDS),
            ),
        )
    finally:
        wb.close()


def read_business_rows(file_path: str | Path, column_mapping: ColumnMapping | None = None) -> list[BusinessRow]:
    inspection = inspect_business_workbook(file_path)
    mapping = column_mapping or inspection.suggested_mapping

    missing = []
    if mapping.sku_col is None:
        missing.append("SKU列（关键词：sku / skuID / 上播 SKU ID，或手动选择）")
    if mapping.code_col is None:
        missing.append("绑定值列（关键词：券码 / 价码 / 促销编码 / 专享券 / 专享价 / 达人id，或手动选择）")
    if missing:
        header_text = "、".join(column.header for column in inspection.columns if column.header)
        raise ValueError(f"未找到必需列: {', '.join(missing)}；当前表头: {header_text or '空'}")

    wb = load_workbook(file_path, data_only=True, read_only=True)
    try:
        ws = wb.active
        sku_col = _require_column_index(mapping.sku_col, ws.max_column, "SKU列")
        code_col = _require_column_index(mapping.code_col, ws.max_column, "绑定值列")
        product_name_col = _optional_column_index(mapping.product_name_col, ws.max_column, "商品名称列")
        selling_point_col = _optional_column_index(mapping.selling_point_col, ws.max_column, "短卖点列")

        rows: list[BusinessRow] = []
        for row_index in range(2, ws.max_row + 1):
            sku = format_cell_value(ws.cell(row_index, sku_col).value)
            if not sku:
                continue
            raw_code = format_cell_value(ws.cell(row_index, code_col).value)
            product_name = format_cell_value(ws.cell(row_index, product_name_col).value) if product_name_col else ""
            selling_point = format_cell_value(ws.cell(row_index, selling_point_col).value) if selling_point_col else ""
            rows.append(
                BusinessRow(
                    source_row=row_index,
                    sku=sku,
                    raw_code=raw_code,
                    product_name=product_name,
                    selling_point=selling_point,
                )
            )

        return rows
    finally:
        wb.close()


def _require_column_index(value: int | None, max_column: int, label: str) -> int:
    if value is None:
        raise ValueError(f"请选择{label}")
    try:
        column_index = int(value)
    except (TypeError, ValueError):
        raise ValueError(f"{label}不是有效列")
    if column_index < 1 or column_index > max_column:
        raise ValueError(f"{label}超出当前表格列范围")
    return column_index


def _optional_column_index(value: int | None, max_column: int, label: str) -> int | None:
    if value in (None, ""):
        return None
    return _require_column_index(value, max_column, label)
