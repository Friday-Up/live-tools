from __future__ import annotations

"""Write promotion binding generation reports."""

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from promotion_binding.template_writer import BindingRecord


@dataclass(frozen=True)
class IssueRecord:
    source_row: int
    sku: str
    raw_code: str
    issue_type: str
    message: str
    action: str
    product_name: str = ""
    kept_source_row: int | None = None


MANUAL_ISSUE_TYPES = {
    "INVALID_CODE",
    "MULTIPLE_KEYS",
    "MULTIPLE_PROMO_IDS",
    "KEY_PROMO_CONFLICT",
    "SKU_BINDING_CONFLICT",
}

SKIPPED_ISSUE_TYPES = {"EMPTY_CODE", "DUPLICATE_BINDING"}

ISSUE_LABELS = {
    "EMPTY_CODE": "未填写券码/价码",
    "INVALID_CODE": "未识别到有效券码/促销ID",
    "MULTIPLE_KEYS": "同一单元格有多个券码 KEY",
    "MULTIPLE_PROMO_IDS": "同一单元格有多个促销 ID",
    "KEY_PROMO_CONFLICT": "同一单元格同时有券码 KEY 和促销 ID",
    "SKU_BINDING_CONFLICT": "同一 SKU 出现多个不同绑定值",
    "DUPLICATE_BINDING": "重复 SKU 和绑定值",
}

BINDING_TYPE_LABELS = {
    "COUPON_KEY": "专享券KEY码",
    "PROMO_ID": "专享价促销ID",
}


def write_report(
    output_path: str | Path,
    summary: dict[str, int],
    binding_records: list[BindingRecord],
    issue_records: list[IssueRecord],
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "汇总"
    detail_ws = wb.create_sheet("可上传明细")
    manual_issue_ws = wb.create_sheet("需处理异常")
    skipped_ws = wb.create_sheet("跳过和重复")

    _write_rows(summary_ws, _summary_rows(summary))
    _write_rows(
        detail_ws,
        [
            ["原始行号", "SKU", "商品名称", "绑定类型", "绑定值", "写入模板行号"],
            *[
                [
                    record.source_row,
                    record.sku,
                    record.product_name,
                    _binding_type_label(record),
                    record.binding_value,
                    template_row,
                ]
                for template_row, record in enumerate(binding_records, start=2)
            ],
        ],
    )
    _write_rows(
        manual_issue_ws,
        [
            ["原始行号", "SKU", "商品名称", "原始券码/价码", "问题", "建议处理"],
            *[
                [
                    issue.source_row,
                    issue.sku,
                    issue.product_name,
                    issue.raw_code,
                    _issue_label(issue),
                    issue.action,
                ]
                for issue in issue_records
                if issue.issue_type in MANUAL_ISSUE_TYPES
            ],
        ],
    )
    _write_rows(
        skipped_ws,
        [
            ["原始行号", "SKU", "商品名称", "原始券码/价码", "跳过原因", "保留行号", "说明"],
            *[
                [
                    issue.source_row,
                    issue.sku,
                    issue.product_name,
                    issue.raw_code,
                    _issue_label(issue),
                    issue.kept_source_row,
                    issue.action,
                ]
                for issue in issue_records
                if issue.issue_type in SKIPPED_ISSUE_TYPES
            ],
        ],
    )

    for ws in wb.worksheets:
        _format_sheet(ws)

    wb.save(output_path)
    return output_path


def _write_rows(ws, rows):
    for row in rows:
        ws.append(list(row))


def _summary_rows(summary: dict[str, int]) -> list[list[object]]:
    categories = {
        "可绑定条数": ("生成结果", "会写入官方上传模板"),
        "专享券KEY条数": ("生成结果", "写入模板 C 列"),
        "专享价促销ID条数": ("生成结果", "写入模板 D 列"),
        "空值跳过": ("未进入模板", "券码/价码为空，默认不需要上传"),
        "异常条数": ("未进入模板", "需要人工确认后再生成"),
        "重复条数": ("未进入模板", "相同 SKU 和绑定值仅保留第一条"),
    }
    rows: list[list[object]] = [["分类", "指标", "数量", "说明"]]
    for metric, count in summary.items():
        category, description = categories.get(metric, ("其他", ""))
        rows.append([category, metric, count, description])
    return rows


def _binding_type_label(record: BindingRecord) -> str:
    return BINDING_TYPE_LABELS.get(record.binding_type.value, record.binding_type.value)


def _issue_label(issue: IssueRecord) -> str:
    return ISSUE_LABELS.get(issue.issue_type, issue.message)


def _format_sheet(ws):
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col in range(1, ws.max_column + 1):
        values = [ws.cell(row, col).value for row in range(1, ws.max_row + 1)]
        max_len = max((len(str(value)) for value in values if value is not None), default=8)
        ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 10), 60)
