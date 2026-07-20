"""读取红包雨 Excel 并完成列映射。"""

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Tuple, Union

from openpyxl import load_workbook

from . import config
from .models import RedRainResult, RedRainRow


@dataclass
class ColumnMapping:
    activity_name_col: Optional[str] = None
    start_time_col: Optional[str] = None
    end_time_col: Optional[str] = None
    issue_method_col: Optional[str] = None
    red_packet_id_col: Optional[str] = None
    win_probability_col: Optional[str] = None

    def is_valid(self) -> bool:
        return all(
            [
                self.activity_name_col,
                self.start_time_col,
                self.end_time_col,
                self.issue_method_col,
                self.red_packet_id_col,
            ]
        )


def _normalize_header(value) -> str:
    return str(value or "").strip().replace(" ", "").replace("（", "").replace("）", "").replace("*", "")


def _find_column(headers: List[str], aliases: List[str]) -> Optional[str]:
    normalized = [(_normalize_header(header), header) for header in headers]
    normalized_aliases = [_normalize_header(alias) for alias in aliases]
    for value, original in normalized:
        if value in normalized_aliases:
            return original
    for value, original in normalized:
        for alias in normalized_aliases:
            if value and (alias in value or value in alias):
                return original
    return None


def inspect_workbook(file_path: Union[str, Path]) -> Tuple[ColumnMapping, List[str]]:
    workbook = load_workbook(file_path, data_only=True)
    try:
        sheet = workbook.active
        if sheet is None:
            raise ValueError("Excel 中没有工作表")
        headers = [str(cell.value or "").strip() for cell in sheet[1]]
    finally:
        workbook.close()

    return (
        ColumnMapping(
            activity_name_col=_find_column(headers, config.ACTIVITY_NAME_ALIASES),
            start_time_col=_find_column(headers, config.START_TIME_ALIASES),
            end_time_col=_find_column(headers, config.END_TIME_ALIASES),
            issue_method_col=_find_column(headers, config.ISSUE_METHOD_ALIASES),
            red_packet_id_col=_find_column(headers, config.RED_PACKET_ID_ALIASES),
            win_probability_col=_find_column(headers, config.WIN_PROBABILITY_ALIASES),
        ),
        headers,
    )


def _parse_datetime(value, field_name: str) -> datetime:
    if isinstance(value, datetime):
        return value
    text = str(value or "").strip()
    if not text:
        raise ValueError(f"{field_name}为空")
    for fmt in config.DATETIME_INPUT_FORMATS:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    raise ValueError(f"{field_name}格式无法识别: {text}")


def _parse_probability(value) -> Optional[int]:
    if value is None or str(value).strip() == "":
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        raise ValueError(f"中奖概率不是数字: {value}")
    if not number.is_integer():
        raise ValueError(f"中奖概率必须是整数: {value}")
    return int(number)


def read_rows_with_errors(
    file_path: Union[str, Path], mapping: ColumnMapping
) -> Tuple[List[RedRainRow], List[RedRainResult]]:
    if not mapping.is_valid():
        raise ValueError("列映射不完整：必须包含活动名称、开始时间、结束时间、红包发放方式和红包ID")

    workbook = load_workbook(file_path, data_only=True)
    try:
        sheet = workbook.active
        if sheet is None:
            raise ValueError("Excel 中没有工作表")
        headers = [str(cell.value or "").strip() for cell in sheet[1]]
        header_indexes = {header: index for index, header in enumerate(headers)}

        def value_at(values, column_name):
            if not column_name or column_name not in header_indexes:
                return None
            index = header_indexes[column_name]
            return values[index] if index < len(values) else None

        rows = []
        rejected = []
        for row_index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            name_value = value_at(values, mapping.activity_name_col)
            if name_value is None or not str(name_value).strip():
                if all(value is None or str(value).strip() == "" for value in values):
                    continue
            try:
                row = RedRainRow(
                    row_index=row_index,
                    activity_name=str(name_value or "").strip(),
                    start_time=_parse_datetime(value_at(values, mapping.start_time_col), "开始时间"),
                    end_time=_parse_datetime(value_at(values, mapping.end_time_col), "结束时间"),
                    issue_method=str(value_at(values, mapping.issue_method_col) or "").strip(),
                    red_packet_id=str(value_at(values, mapping.red_packet_id_col) or "").strip(),
                    win_probability=_parse_probability(value_at(values, mapping.win_probability_col)),
                )
            except ValueError as exc:
                rejected.append(
                    RedRainResult(
                        row_index=row_index,
                        activity_name=str(name_value or "").strip(),
                        start_time=value_at(values, mapping.start_time_col),
                        end_time=value_at(values, mapping.end_time_col),
                        issue_method=str(value_at(values, mapping.issue_method_col) or "").strip(),
                        red_packet_id=str(value_at(values, mapping.red_packet_id_col) or "").strip(),
                        win_probability=value_at(values, mapping.win_probability_col),
                        status="跳过",
                        error=str(exc),
                    )
                )
                continue
            rows.append(row)
        return rows, rejected
    finally:
        workbook.close()


def read_rows(file_path: Union[str, Path], mapping: ColumnMapping) -> List[RedRainRow]:
    rows, rejected = read_rows_with_errors(file_path, mapping)
    if rejected:
        first = rejected[0]
        raise ValueError(f"第 {first.row_index} 行：{first.error}")
    return rows
