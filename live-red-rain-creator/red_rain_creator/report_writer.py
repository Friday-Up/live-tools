"""生成红包雨批量创建结果 Excel。"""

from datetime import datetime
from pathlib import Path
from typing import Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from . import config
from .models import BatchResult


def write_batch_report(result: BatchResult, output_dir: Union[str, Path]) -> Path:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"红包雨创建结果_{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "创建结果"
    sheet.append(config.RESULT_COLUMNS)
    for item in sorted(result.results, key=lambda value: value.row_index):
        sheet.append(
            [
                item.row_index,
                item.activity_name,
                item.start_time,
                item.end_time,
                item.issue_method,
                item.red_packet_id,
                item.win_probability,
                item.status,
                item.activity_id,
                item.error,
            ]
        )

    header_fill = PatternFill("solid", fgColor="E1251B")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in ("C", "D"):
        for cell in sheet[column][1:]:
            cell.number_format = "yyyy-mm-dd hh:mm:ss"
    widths = {"A": 10, "B": 24, "C": 21, "D": 21, "E": 20, "F": 18, "G": 12, "H": 12, "I": 18, "J": 44}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    workbook.save(output_path)
    workbook.close()
    return output_path
