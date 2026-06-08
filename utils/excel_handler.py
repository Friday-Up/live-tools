"""
Excel 处理模块
负责读取 SKU 列表、写入价格和截图
"""

import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter


def read_sku_list(file_path, sku_column='商品SKU'):
    """
    从 Excel 读取 SKU 列表

    Args:
        file_path: Excel 文件路径
        sku_column: SKU 列的标题名

    Returns:
        list: [(row_index, sku), ...]
    """
    wb = load_workbook(file_path)
    ws = wb.active

    # 找到表头行
    headers = [cell.value for cell in ws[1]]

    # 找到 SKU 列索引
    if sku_column in headers:
        sku_col_index = headers.index(sku_column)
    else:
        # 默认第2列（B列）
        sku_col_index = 1

    sku_list = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        sku = row[sku_col_index]
        if sku:
            sku_list.append((i, str(sku).strip()))

    return sku_list


def write_results(file_path, results, threshold_price, output_dir='output',
                  sku_column='商品SKU', price_column='价格', image_column='图片', remark_column='备注'):
    """
    将测价结果写回 Excel

    Args:
        file_path: 原始 Excel 文件路径
        results: 测价结果列表
        threshold_price: 价格门槛
        output_dir: 输出目录
        price_column: 价格列标题
        image_column: 图片列标题
        remark_column: 备注列标题

    Returns:
        str: 输出文件路径
    """
    wb = load_workbook(file_path)
    ws = wb.active

    # 找到列索引
    headers = [cell.value for cell in ws[1]]

    def get_col_index(col_name):
        if col_name in headers:
            return headers.index(col_name) + 1
        return None

    price_col = get_col_index(price_column) or 4
    img_col = get_col_index(image_column) or 5
    remark_col = get_col_index(remark_column) or 6

    # 设置图片列宽度
    ws.column_dimensions[get_column_letter(img_col)].width = 30

    for result in results:
        row_index = result.get('row_index')
        if not row_index:
            continue

        if result['status'] == 'success':
            # 写入价格
            ws.cell(row=row_index, column=price_col, value=result['price'])

            # 嵌入截图
            if result['screenshot_path'] and os.path.exists(result['screenshot_path']):
                try:
                    img = XLImage(result['screenshot_path'])
                    # 调整图片大小
                    img.width = 200
                    img.height = 150
                    # 设置行高以容纳图片
                    ws.row_dimensions[row_index].height = 120
                    cell_addr = f"{get_column_letter(img_col)}{row_index}"
                    ws.add_image(img, cell_addr)
                except Exception as e:
                    print(f"  ⚠️ 嵌入图片失败: {e}")

            # 标记不合格
            if result['price'] is not None and result['price'] < threshold_price:
                ws.cell(row=row_index, column=remark_col, value="不符合上菜")

        elif result['status'] == 'need_login':
            ws.cell(row=row_index, column=remark_col,
                   value=f"需要登录: {result['message']}")

        else:
            ws.cell(row=row_index, column=remark_col,
                   value=f"抓取失败: {result['message']}")

    # 保存到 output 目录
    os.makedirs(output_dir, exist_ok=True)
    base_name = os.path.basename(file_path)
    output_name = base_name.replace('.xlsx', '_result.xlsx')
    output_path = os.path.join(output_dir, output_name)
    wb.save(output_path)
    print(f"\n✅ 结果已保存: {output_path}")
    return output_path
