import tempfile
import unittest
from collections import defaultdict
from pathlib import Path
from unittest.mock import MagicMock, patch

from openpyxl import Workbook

from utils.excel_handler import read_sku_list, write_results, create_sku_input_file


def make_workbook(headers, rows):
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(tmp.name)
    wb.close()
    return Path(tmp.name)


class ExcelHandlerTests(unittest.TestCase):
    def test_read_sku_list_closes_workbook_after_reading(self):
        workbook = MagicMock()
        worksheet = MagicMock()
        workbook.active = worksheet
        worksheet.__getitem__.return_value = [MagicMock(value="商品SKU")]
        worksheet.iter_rows.return_value = [("100264886683",)]

        with patch("utils.excel_handler.load_workbook", return_value=workbook):
            self.assertEqual(read_sku_list("input.xlsx", "商品SKU"), [(2, "100264886683")])

        workbook.close.assert_called_once()

    def test_write_results_closes_workbook_after_saving(self):
        workbook = MagicMock()
        worksheet = MagicMock()
        workbook.active = worksheet
        worksheet.__getitem__.return_value = [MagicMock(value="商品SKU")]
        worksheet.max_column = 1
        worksheet.cell.return_value.value = None
        worksheet.column_dimensions = defaultdict(MagicMock)
        worksheet.row_dimensions = defaultdict(MagicMock)

        with patch("utils.excel_handler.load_workbook", return_value=workbook):
            write_results(
                file_path="input.xlsx",
                results=[],
                threshold_price=6,
                output_dir=tempfile.mkdtemp(),
            )

        workbook.save.assert_called_once()
        workbook.close.assert_called_once()

    def test_reads_required_sku_header_with_suffix(self):
        path = make_workbook(
            ["提交时间（自动）", "商品SKU（必填）", "提交者（自动）"],
            [["2026-06-11", "100264886683", "张三"]],
        )

        self.assertEqual(read_sku_list(str(path), "商品SKU"), [(2, "100264886683")])

    def test_rejects_missing_sku_header_instead_of_guessing_column_b(self):
        path = make_workbook(
            ["提交时间", "备注"],
            [["2026-06-11", "这不是SKU"]],
        )

        with self.assertRaisesRegex(ValueError, "未找到SKU列"):
            read_sku_list(str(path), "商品SKU")

    def test_write_results_adds_missing_output_headers(self):
        path = make_workbook(
            ["提交时间（自动）", "商品SKU（必填）", "提交者（自动）"],
            [["2026-06-11", "100264886683", "张三"]],
        )
        out_dir = tempfile.mkdtemp()

        output = write_results(
            file_path=str(path),
            results=[
                {
                    "row_index": 2,
                    "sku": "100264886683",
                    "price": 5.5,
                    "screenshot_path": None,
                    "status": "success",
                    "message": "ok",
                }
            ],
            threshold_price=6,
            output_dir=out_dir,
        )

        from openpyxl import load_workbook

        wb = load_workbook(output)
        try:
            ws = wb.active
            self.assertEqual(
                [ws.cell(1, i).value for i in range(1, 7)],
                ["提交时间（自动）", "商品SKU（必填）", "提交者（自动）", "价格", "图片", "备注"],
            )
            self.assertEqual(ws.cell(2, 4).value, 5.5)
            self.assertEqual(ws.cell(2, 6).value, "不符合上菜")
        finally:
            wb.close()

    def test_write_results_appends_output_headers_when_fallback_columns_are_occupied(self):
        path = make_workbook(
            ["提交时间", "商品SKU", "提交者", "活动名称"],
            [["2026-06-11", "100264886683", "张三", "618直播"]],
        )
        out_dir = tempfile.mkdtemp()

        output = write_results(
            file_path=str(path),
            results=[
                {
                    "row_index": 2,
                    "sku": "100264886683",
                    "price": 5.5,
                    "screenshot_path": None,
                    "status": "success",
                    "message": "ok",
                }
            ],
            threshold_price=6,
            output_dir=out_dir,
        )

        from openpyxl import load_workbook

        wb = load_workbook(output)
        try:
            ws = wb.active
            self.assertEqual(ws.cell(1, 4).value, "活动名称")
            self.assertEqual([ws.cell(1, i).value for i in range(5, 8)], ["价格", "图片", "备注"])
            self.assertEqual(ws.cell(2, 5).value, 5.5)
            self.assertEqual(ws.cell(2, 7).value, "不符合上菜")
        finally:
            wb.close()

    def test_write_results_records_partial_price_and_manual_review_remark(self):
        path = make_workbook(
            ["提交时间", "商品SKU", "提交者"],
            [["2026-06-11", "48279162646", "张三"]],
        )
        out_dir = tempfile.mkdtemp()

        output = write_results(
            file_path=str(path),
            results=[
                {
                    "row_index": 2,
                    "sku": "48279162646",
                    "price": 8.0,
                    "screenshot_path": None,
                    "status": "partial",
                    "message": "需人工复核: 部分系列/规格未完成检测；已检测 1 个规格，最低 ¥8.0",
                }
            ],
            threshold_price=6,
            output_dir=out_dir,
        )

        from openpyxl import load_workbook

        wb = load_workbook(output)
        try:
            ws = wb.active
            self.assertEqual(ws.cell(2, 4).value, 8.0)
            self.assertEqual(
                ws.cell(2, 6).value,
                "需人工复核: 部分系列/规格未完成检测；已检测 1 个规格，最低 ¥8.0",
            )
        finally:
            wb.close()

    def test_create_sku_input_file_generates_readable_sku_list(self):
        out_dir = tempfile.mkdtemp()
        output_path = Path(out_dir) / 'sku_input.xlsx'
        create_sku_input_file(['100264886683', '48279162646'], str(output_path))
        self.assertTrue(output_path.exists())
        self.assertEqual(
            read_sku_list(str(output_path), '商品SKU'),
            [(2, '100264886683'), (3, '48279162646')],
        )


if __name__ == "__main__":
    unittest.main()
