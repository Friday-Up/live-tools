"""京东多来源选品 Agent 服务编排与报表输出。"""
from __future__ import annotations

import datetime as dt
import json
import os
from dataclasses import dataclass
from pathlib import Path

from . import config
from .parser import parse_all
from .recommender import recommend
from .runtime import RunContext
from .selector import build_candidate_pool


def _timestamp() -> str:
    return dt.datetime.now().strftime("%Y%m%d_%H%M%S")


def _final_selection(recommendation: dict) -> dict:
    return {
        source_name: {
            category_name: block.get("products", [])
            for category_name, block in categories.items()
        }
        for source_name, categories in recommendation.items()
    }


def _annotate_candidate_pool(candidate_pool: dict, recommendation: dict) -> dict:
    annotated: dict = {}
    for source_name, categories in candidate_pool.items():
        annotated[source_name] = {}
        for category_name, products in categories.items():
            block = recommendation[source_name][category_name]
            decisions = block.get("candidate_decisions", {})
            rows = []
            for product in products:
                row = dict(product)
                decision = decisions.get(product["sku_id"], {})
                row["final_selected"] = bool(decision.get("selected"))
                row["ai_selected"] = (
                    row["final_selected"]
                    and block.get("recommendation_mode") == "llm_enhanced"
                )
                row["ai_rank"] = decision.get("ai_rank")
                row["rejection_reason"] = decision.get("rejection_reason", "")
                row["reco_score"] = decision.get("reco_score")
                row["score_detail"] = decision.get("score_detail", {})
                row["recommendation_mode"] = block.get("recommendation_mode")
                row["shortfall_reason"] = block.get("shortfall_reason", "")
                rows.append(row)
            annotated[source_name][category_name] = rows
    return annotated


def _diagnostics(
    raw_data: dict,
    candidate_pool: dict,
    selection: dict,
    recommendation: dict,
) -> dict:
    sources = {}
    for source_key, payload in raw_data.items():
        source_name = payload.get("name", source_key)
        candidate_categories = candidate_pool.get(source_name, {})
        categories = selection.get(source_name, {})
        sources[source_key] = {
            "name": source_name,
            "adapter": payload.get("adapter", "offline"),
            "status": payload.get("status", "ok" if payload.get("goods") else "empty"),
            "error": payload.get("error", ""),
            "raw_goods_count": len(payload.get("goods", [])),
            "category_count": len(candidate_categories),
            "candidate_goods_count": sum(len(goods) for goods in candidate_categories.values()),
            "selected_goods_count": sum(len(goods) for goods in categories.values()),
            "short_categories": {
                name: {
                    "selected_count": len(goods),
                    "reason": recommendation[source_name][name].get("shortfall_reason")
                    or "最终入选不足 10 个",
                }
                for name, goods in categories.items()
                if len(goods) < config.TOP_N_PER_CATEGORY
            },
        }
    missing = [item["name"] for item in sources.values() if item["status"] != "ok"]
    return {"complete": not missing, "missing_sources": missing, "sources": sources}


def _ai_diagnostics(recommendation: dict) -> dict:
    failed = []
    category_count = 0
    for source_name, categories in recommendation.items():
        for category_name, block in categories.items():
            category_count += 1
            if block.get("recommendation_mode") == "llm_enhanced" and not block.get("ai_error"):
                continue
            failed.append(
                {
                    "source": source_name,
                    "category": category_name,
                    "error": block.get("ai_error")
                    or f"推荐模式为 {block.get('recommendation_mode', 'unknown')}",
                }
            )
    return {
        "ai_complete": category_count > 0 and not failed,
        "ai_failed_categories": failed,
    }


def _load_offline(path: str) -> dict:
    with open(path, encoding="utf-8") as handle:
        raw = json.load(handle)
    if isinstance(raw, list):
        from product_selection_agent.fetcher import _extract_goods_from_body

        goods = []
        for body in raw:
            goods.extend(_extract_goods_from_body(body))
        return {"gov_subsidy": {"name": "国家补贴", "adapter": "offline", "goods": goods}}
    if isinstance(raw, dict) and "selection" in raw and "items_count" in raw:
        raise ValueError("--offline 需要原始抓取 JSON，不是已经生成的 selection 结果")
    if not isinstance(raw, dict):
        raise ValueError("--offline 文件必须是对象或 qryJediPcBabelFloors 响应数组")
    return raw


def run_selection(
    headless: bool = False,
    offline_path: str | None = None,
    allow_partial: bool = False,
    context: RunContext | None = None,
) -> dict:
    context = context or RunContext()
    context.check_cancelled()
    if offline_path:
        raw_data = _load_offline(offline_path)
    else:
        from .fetcher import fetch_all

        raw_data = fetch_all(
            headless=headless,
            allow_partial=allow_partial,
            context=context,
        )

    context.check_cancelled()
    items = parse_all(raw_data)
    context.log(f"[main] 解析有效商品 {len(items)} 个")
    candidate_pool = build_candidate_pool(items)
    for source_name, categories in candidate_pool.items():
        for category_name, products in categories.items():
            context.log(f"[candidate] {source_name}/{category_name}: 候选 {len(products)} 个")
    recommendation = recommend(candidate_pool, context=context)
    selection = _final_selection(recommendation)
    candidate_pool = _annotate_candidate_pool(candidate_pool, recommendation)
    for source_name, categories in selection.items():
        for category_name, products in categories.items():
            context.log(f"[select] {source_name}/{category_name}: 最终入选 {len(products)} 个")
    diagnostics = _diagnostics(raw_data, candidate_pool, selection, recommendation)
    diagnostics["fetch_complete"] = diagnostics["complete"]
    diagnostics.update(_ai_diagnostics(recommendation))

    modes = {
        data.get("recommendation_mode", "explainable_scoring")
        for categories in recommendation.values()
        for data in categories.values()
    }
    mode = modes.pop() if len(modes) == 1 else ("mixed" if modes else "explainable_scoring")
    total_categories = sum(len(categories) for categories in selection.values())
    context.log(
        f"[main] 覆盖来源 {len(selection)} 个，类目 {total_categories} 个，"
        f"抓取完整={diagnostics['fetch_complete']}，AI完整={diagnostics['ai_complete']}"
    )
    return {
        "generated_at": dt.datetime.now().astimezone().isoformat(timespec="seconds"),
        "contract": "每来源 × 页面类目：AI 从最多30个候选中筛选最多Top10；合格不足不补齐",
        "items_count": len(items),
        "recommendation_mode": mode,
        "ai_complete": diagnostics["ai_complete"],
        "diagnostics": diagnostics,
        "candidate_pool": candidate_pool,
        "selection": selection,
        "recommendation": recommendation,
    }


def save_json(payload: dict, out_dir: str, timestamp: str) -> str:
    path = os.path.join(out_dir, f"selection_{timestamp}.json")
    with open(path, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)
    return path


def _autosize_and_filter(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column in worksheet.columns:
        width = max((len(str(cell.value)) for cell in column if cell.value is not None), default=10)
        worksheet.column_dimensions[column[0].column_letter].width = min(width + 2, 55)


def save_excel(payload: dict, out_dir: str, timestamp: str) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    candidates = workbook.active
    candidates.title = "候选池"
    candidates.append([
        "来源", "页面类目", "候选排名", "最终是否入选", "AI是否入选", "AI推荐顺位",
        "淘汰原因", "合格不足说明", "推荐模式", "规则参考分", "评分明细", "商品名", "SKU",
        "页面/榜单位次", "榜单名", "到手价", "京东价", "划线价", "销量数",
        "销量口径", "折扣力度", "好评率", "自营", "卖点", "店铺", "商品链接",
    ])
    for source_name, categories in payload["candidate_pool"].items():
        for category_name, products in categories.items():
            for product in products:
                candidates.append([
                    source_name, category_name, product.get("candidate_rank"),
                    "是" if product.get("final_selected") else "否",
                    "是" if product.get("ai_selected") else "否", product.get("ai_rank"),
                    product.get("rejection_reason"), product.get("shortfall_reason"),
                    product.get("recommendation_mode"), product.get("reco_score"),
                    json.dumps(product.get("score_detail", {}), ensure_ascii=False),
                    product.get("name"), product.get("sku_id"), product.get("source_rank"),
                    product.get("rank_board"), product.get("display_price"),
                    product.get("jd_price"), product.get("line_price"), product.get("sales_num"),
                    product.get("sales_text"), product.get("discount_ratio"),
                    product.get("good_rate"), "是" if product.get("jx_self") else "否",
                    product.get("selling_points"), product.get("shop_name"), product.get("url"),
                ])

    detail = workbook.create_sheet("选品明细")
    detail.append([
        "来源", "页面类目", "选品排名", "页面/榜单位次", "榜单名", "商品名", "SKU",
        "到手价", "京东价", "划线价", "销量数", "销量口径", "折扣力度", "好评率",
        "自营", "卖点", "类目来源", "数据说明", "商品链接", "榜单链接",
    ])
    for source_name, categories in payload["selection"].items():
        for category_name, products in categories.items():
            for product in products:
                detail.append([
                    source_name, category_name, product.get("rank"), product.get("source_rank"),
                    product.get("rank_board"), product.get("name"), product.get("sku_id"),
                    product.get("display_price"), product.get("jd_price"), product.get("line_price"),
                    product.get("sales_num"), product.get("sales_text"), product.get("discount_ratio"),
                    product.get("good_rate"), "是" if product.get("jx_self") else "否",
                    product.get("selling_points"), product.get("category_source"),
                    product.get("data_quality"), product.get("url"), product.get("rank_board_url"),
                ])

    reco = workbook.create_sheet("推荐结果")
    reco.append([
        "来源", "页面类目", "推荐顺位", "商品名", "SKU", "到手价", "销量口径",
        "推荐分", "评分明细", "推荐模式", "推荐理由", "推荐文案", "商品链接",
    ])
    for source_name, categories in payload["recommendation"].items():
        for category_name, data in categories.items():
            for index, product in enumerate(data.get("products", []), 1):
                reco.append([
                    source_name, category_name, index, product.get("name"), product.get("sku_id"),
                    product.get("display_price"), product.get("sales_text"), product.get("reco_score"),
                    json.dumps(product.get("score_detail", {}), ensure_ascii=False),
                    product.get("recommendation_mode"), product.get("reason"), product.get("copy"),
                    product.get("url"),
                ])

    diag = workbook.create_sheet("运行诊断")
    diag.append([
        "来源", "适配器", "状态", "原始商品数", "候选商品数", "类目数",
        "入选商品数", "不足10条类目及原因", "错误",
    ])
    for item in payload["diagnostics"]["sources"].values():
        diag.append([
            item["name"], item["adapter"], item["status"], item["raw_goods_count"],
            item.get("candidate_goods_count", 0), item["category_count"], item["selected_goods_count"],
            json.dumps(item["short_categories"], ensure_ascii=False), item["error"],
        ])

    for worksheet in (candidates, detail, reco, diag):
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="C00000")
        _autosize_and_filter(worksheet)

    path = os.path.join(out_dir, f"selection_{timestamp}.xlsx")
    workbook.save(path)
    return path


@dataclass(frozen=True)
class SelectionRunResult:
    payload: dict
    json_path: Path
    excel_path: Path


def execute_selection(
    output_dir: str | Path,
    headless: bool = True,
    offline_path: str | None = None,
    allow_partial: bool = False,
    context: RunContext | None = None,
) -> SelectionRunResult:
    """执行完整选品并将两个审计结果写入指定目录。"""
    context = context or RunContext()
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    payload = run_selection(
        headless=headless,
        offline_path=offline_path,
        allow_partial=allow_partial,
        context=context,
    )
    context.check_cancelled()
    timestamp = _timestamp()
    json_path = Path(save_json(payload, str(output_dir), timestamp))
    excel_path = Path(save_excel(payload, str(output_dir), timestamp))
    context.log(f"[main] JSON: {json_path}")
    context.log(f"[main] Excel: {excel_path}")
    context.log(f"[main] 推荐模式: {payload['recommendation_mode']}")
    return SelectionRunResult(payload=payload, json_path=json_path, excel_path=excel_path)
