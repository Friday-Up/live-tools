import json
import os
import sys
import tempfile
import time
import unittest
from unittest import mock


PROJECT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_DIR)

from product_selection_agent import fetcher
from product_selection_agent.fetcher import (
    _dedup_goods,
    _extract_flex_goods,
    _extract_goods_from_body,
    _limit_candidates,
    _rank_product_to_goods,
)
from product_selection_agent.parser import parse_all, parse_quantity_text
from product_selection_agent import recommender
from product_selection_agent.recommender import recommend
from product_selection_agent.recommender import _extract_json_object
from product_selection_agent.recommender import _read_stream_content
from product_selection_agent.runtime import RunContext
from product_selection_agent.selector import select_top


class QuantityParserTest(unittest.TestCase):
    def test_chinese_units(self):
        self.assertEqual(parse_quantity_text("30日售出1万件"), 10000)
        self.assertEqual(parse_quantity_text("2.3万人买过"), 23000)
        self.assertEqual(parse_quantity_text("已售5000+"), 5000)
        self.assertEqual(parse_quantity_text("无销量"), 0)


class SourceAdapterTest(unittest.TestCase):
    def test_candidate_pool_is_capped(self):
        goods = [{"skuId": str(index), "tab_category": "测试"} for index in range(40)]
        with mock.patch.object(fetcher.config, "MAX_CANDIDATES_PER_CATEGORY", 30):
            self.assertEqual(len(_limit_candidates(goods)), 30)

    def test_sources_use_isolated_workers_concurrently(self):
        sources = [
            {"key": f"s{index}", "name": f"来源{index}", "adapter": "test"}
            for index in range(4)
        ]

        def fake_fetch(
            source,
            headless,
            auth_path,
            context=None,
            selected_categories=None,
        ):
            time.sleep(0.2)
            return [{"skuId": source["key"], "tab_category": "类目"}]

        start = time.monotonic()
        with mock.patch.multiple(
            fetcher.config,
            SOURCES=sources,
            FETCH_WORKERS=4,
            AUTH_PATH="/path/does/not/exist",
        ), mock.patch("product_selection_agent.fetcher._fetch_source_isolated", side_effect=fake_fetch):
            result = fetcher.fetch_all(headless=True)
        elapsed = time.monotonic() - start

        self.assertEqual(list(result), [source["key"] for source in sources])
        self.assertLess(elapsed, 0.6)

    def test_isolated_browser_keeps_run_context_for_logging_and_stop(self):
        source = {"key": "test", "name": "测试来源", "adapter": "test"}
        run_context = RunContext(log_callback=lambda _message: None)
        page = mock.MagicMock(name="page")
        browser_context = mock.MagicMock(name="browser_context")
        browser_context.new_page.return_value = page
        browser = mock.MagicMock(name="browser")
        browser.new_context.return_value = browser_context
        playwright = mock.MagicMock(name="playwright")
        playwright.chromium.launch.return_value = browser
        manager = mock.MagicMock(name="playwright_manager")
        manager.__enter__.return_value = playwright

        with mock.patch.object(fetcher, "sync_playwright", return_value=manager), mock.patch.object(
            fetcher, "fetch_source", return_value=[]
        ) as fetch_source:
            fetcher._fetch_source_isolated(
                source,
                headless=True,
                auth_path="/path/does/not/exist",
                context=run_context,
            )

        fetch_source.assert_called_once_with(page, source, run_context, None)

    def test_selected_categories_skip_unselected_sources_before_browser_start(self):
        sources = [
            {"key": "a", "name": "来源A", "adapter": "test"},
            {"key": "b", "name": "来源B", "adapter": "test"},
        ]
        calls = []

        def fake_fetch(
            source,
            headless,
            auth_path,
            context=None,
            selected_categories=None,
        ):
            calls.append((source["key"], selected_categories))
            return [{"skuId": "1", "tab_category": next(iter(selected_categories))}]

        with mock.patch.multiple(
            fetcher.config,
            SOURCES=sources,
            FETCH_WORKERS=2,
            AUTH_PATH="/path/does/not/exist",
        ), mock.patch(
            "product_selection_agent.fetcher._fetch_source_isolated",
            side_effect=fake_fetch,
        ):
            result = fetcher.fetch_all(
                headless=True,
                selected_categories={"a": ["类目A"], "b": []},
            )

        self.assertEqual(list(result), ["a"])
        self.assertEqual(calls, [("a", {"类目A"})])

    def test_tab_plan_only_keeps_selected_page_categories(self):
        plan, unknown = fetcher._build_tab_plan(
            ["推荐", "手机", "电视", "空调"],
            {"电视", "空调"},
        )

        self.assertEqual(plan, [(2, "电视"), (3, "空调")])
        self.assertEqual(unknown, set())

    def test_discover_categories_keeps_source_order_and_errors(self):
        sources = [
            {"key": "a", "name": "来源A", "adapter": "test"},
            {"key": "b", "name": "来源B", "adapter": "test"},
        ]

        def fake_discover(source, headless, auth_path, context=None):
            if source["key"] == "b":
                raise RuntimeError("页面变化")
            return ["手机", "电视"]

        with mock.patch.multiple(
            fetcher.config,
            SOURCES=sources,
            FETCH_WORKERS=2,
            AUTH_PATH="/path/does/not/exist",
        ), mock.patch(
            "product_selection_agent.fetcher._discover_source_categories_isolated",
            side_effect=fake_discover,
        ):
            result = fetcher.discover_categories(headless=True)

        self.assertEqual(list(result), ["a", "b"])
        self.assertEqual(result["a"]["categories"], ["手机", "电视"])
        self.assertEqual(result["b"]["categories"], [])
        self.assertIn("页面变化", result["b"]["error"])

    def test_gov_goods_nested_sku_is_kept(self):
        body = {
            "floorResponse": {
                "floor": {
                    "providerData": {
                        "feeds": {
                            "goodsList": [
                                {
                                    "jump": {"params": {"skuId": "nested-1"}},
                                    "wname": "嵌套 SKU 商品",
                                    "tab_category": "测试类目",
                                }
                            ]
                        }
                    }
                }
            }
        }

        goods = _extract_goods_from_body(body)
        parsed = parse_all({"gov_subsidy": {"name": "国家补贴", "goods": goods}})

        self.assertEqual(len(goods), 1)
        self.assertEqual(parsed[0]["sku_id"], "nested-1")

    def test_special_feed_sample(self):
        groups = []
        for index in range(10):
            sku_id = "10148311755369" if index == 4 else str(10120924581691 + index)
            groups.append(
                {
                    "clickEvent": {"jump": {"params": {"skuId": sku_id}}},
                    "flexData": {
                        "wname": f"特价商品{index + 1}",
                        "displayPrice": str(10 + index),
                        "saleQttyWholeDesc1": "已售30万+" if index == 4 else "已售100+",
                    },
                }
            )
        body = {"result": {"groups": groups}}
        goods = _extract_flex_goods(body, "推荐")
        self.assertEqual(len(goods), 10)
        self.assertEqual(goods[0]["skuId"], "10120924581691")
        self.assertTrue(goods[0]["wname"])
        self.assertGreater(float(goods[0]["displayPrice"]), 0)
        hot = next(item for item in goods if item["skuId"] == "10148311755369")
        parsed = parse_all({"jd_special": {"name": "京东特价", "goods": [hot]}})[0]
        self.assertEqual(parsed["sales_num"], 300000)

    def test_rank_product_normalization(self):
        raw = _rank_product_to_goods(
            {
                "skuId": "123",
                "name": "测试商品",
                "rankNum": 2,
                "threeCategory": "1105",
                "storeId": "10",
                "zyTag": "1",
                "price": {"purchasePrice": "99", "jdPrice": "129"},
                "skuBenefitTags": [{"text": "30日售出1万件"}],
                "skuInfoTags": [{"ext": {"sellPoints": ["高刷新率"]}}],
            },
            "3C数码",
            "电脑热卖榜",
            "https://example.test/rank",
        )
        parsed = parse_all({"rank_list": {"name": "排行榜", "goods": [raw]}})[0]
        self.assertEqual(parsed["sales_num"], 10000)
        self.assertEqual(parsed["source_rank"], 2)
        self.assertEqual(parsed["rank_board"], "电脑热卖榜")
        self.assertTrue(parsed["jx_self"])


class SelectionAndRecommendationTest(unittest.TestCase):
    def test_selection_prompt_uses_commercial_metrics_only_for_ranking(self):
        from product_selection_agent.recommender import _build_selection_prompt

        prompt = _build_selection_prompt(
            "国家补贴",
            "医疗器械",
            [{"sku_id": "1", "name": "医用血压计", "sales_num": 10}],
            10,
        )

        self.assertIn("低销量、高价格、折扣弱、缺少好评率、非自营", prompt)
        self.assertIn("只能影响排序，不能作为淘汰理由", prompt)
        self.assertIn("相关有效商品达到10个时必须选择10个", prompt)
        self.assertIn("血压计、电动轮椅属于医疗器械", prompt)
        self.assertIn("黑芝麻丸属于食品", prompt)
        self.assertIn("rejected 只列出被硬过滤的候选", prompt)
        self.assertIn('"selected_sku_ids":["..."]', prompt)
        self.assertIn("selected_sku_ids 中的顺序就是推荐排名", prompt)
        self.assertIn("selected_reasons 可选", prompt)
        self.assertNotIn('"selected":[{"sku_id"', prompt)

    def test_ordered_sku_ids_use_program_explanations_when_optional_details_missing(self):
        from product_selection_agent.recommender import _normalize_ai_selection

        candidates = [
            {
                "sku_id": "1",
                "name": "商品一",
                "display_price": 99,
                "sales_text": "已售5000+",
                "source_rank": 1,
            },
            {
                "sku_id": "2",
                "name": "商品二",
                "display_price": 79,
                "sales_text": "已售3000+",
                "source_rank": 2,
            },
        ]
        parsed = {
            "selected_sku_ids": ["2", "1"],
            "rejected": [],
            "shortfall_reason": "仅有 2 个候选商品",
        }

        normalized = _normalize_ai_selection(parsed, candidates, limit=10)

        self.assertEqual(
            [item["sku_id"] for item in normalized["selected"]],
            ["2", "1"],
        )
        self.assertEqual([item["rank"] for item in normalized["selected"]], [1, 2])
        self.assertIn("当前展示价 79 元", normalized["selected"][0]["reason"])
        self.assertIn("商品二", normalized["selected"][0]["copy"])
        self.assertTrue(normalized["protocol_complete"])

    def test_optional_ai_reason_overrides_program_generated_explanation(self):
        from product_selection_agent.recommender import _normalize_ai_selection

        normalized = _normalize_ai_selection(
            {
                "selected_sku_ids": ["1"],
                "selected_reasons": [
                    {"sku_id": "1", "reason": "AI 推荐理由", "copy": "AI 推荐文案"}
                ],
                "rejected": [],
                "shortfall_reason": "仅有 1 个候选商品",
            },
            [{"sku_id": "1", "name": "商品一", "display_price": 99}],
            limit=10,
        )

        self.assertEqual(normalized["selected"][0]["reason"], "AI 推荐理由")
        self.assertEqual(normalized["selected"][0]["copy"], "AI 推荐文案")

    def test_invalid_ordered_skus_make_ai_protocol_incomplete(self):
        from product_selection_agent.recommender import _normalize_ai_selection

        normalized = _normalize_ai_selection(
            {
                "selected_sku_ids": ["1", "1", "unknown"],
                "rejected": [],
                "shortfall_reason": "相关有效商品达到10个，已选择10个",
            },
            [{"sku_id": str(index), "name": f"商品{index}"} for index in range(1, 11)],
            limit=10,
        )

        self.assertFalse(normalized["protocol_complete"])
        self.assertIn("重复或未知 SKU", normalized["protocol_error"])
        self.assertIn("实际入选 1 个", normalized["protocol_error"])

    def test_legacy_selected_without_reason_or_copy_keeps_selected_sku(self):
        from product_selection_agent.recommender import (
            _coerce_ai_selection_payload,
            _normalize_ai_selection,
        )

        coerced = _coerce_ai_selection_payload(
            {
                "selected": [{"sku_id": "1", "rank": 1}],
                "rejected": [],
                "shortfall_reason": "仅有 1 个候选商品",
            }
        )
        normalized = _normalize_ai_selection(
            coerced,
            [{"sku_id": "1", "name": "商品一", "display_price": 99}],
            limit=10,
        )

        self.assertEqual([item["sku_id"] for item in normalized["selected"]], ["1"])
        self.assertIn("当前展示价 99 元", normalized["selected"][0]["reason"])

    def test_legacy_items_without_reason_or_copy_keep_selected_skus(self):
        from product_selection_agent.recommender import (
            _coerce_ai_selection_payload,
            _normalize_ai_selection,
        )

        coerced = _coerce_ai_selection_payload(
            {"items": [{"sku_id": "2"}, {"sku_id": "1"}]}
        )
        normalized = _normalize_ai_selection(
            coerced,
            [
                {"sku_id": "1", "name": "商品一", "display_price": 99},
                {"sku_id": "2", "name": "商品二", "display_price": 79},
            ],
            limit=2,
        )

        self.assertEqual(
            [item["sku_id"] for item in normalized["selected"]],
            ["2", "1"],
        )
        self.assertTrue(normalized["protocol_complete"])

    def test_incomplete_ai_protocol_falls_back_to_rule_selection(self):
        candidates = [
            {
                "sku_id": str(index),
                "name": f"商品{index}",
                "display_price": 10 + index,
                "sales_num": 100 - index,
                "source_rank": index,
            }
            for index in range(1, 11)
        ]
        invalid_decision = {
            "selected": [],
            "rejected": {},
            "shortfall_reason": "相关有效商品达到10个，已选择10个",
            "warnings": ["忽略重复 SKU"],
            "protocol_complete": False,
            "protocol_error": "返回含重复 SKU，实际入选 8 个",
        }

        with mock.patch.multiple(
            "product_selection_agent.recommender.config",
            AI_API_URL="http://model.test/chat/completions",
            AI_API_KEY="key",
            AI_MODEL="model",
        ), mock.patch(
            "product_selection_agent.recommender._llm_select_category",
            return_value=invalid_decision,
        ):
            block = recommend({"测试来源": {"测试类目": candidates}})["测试来源"]["测试类目"]

        self.assertEqual(block["recommendation_mode"], "explainable_scoring")
        self.assertEqual(len(block["products"]), 10)
        self.assertIn("AI 选品协议不完整", block["ai_error"])

    def test_platform_category_ids_filter_known_cross_category_products(self):
        from product_selection_agent.recommender import _prefilter_candidates

        medical = [
            {"sku_id": "food", "category_id2": "9195"},
            {"sku_id": "device", "category_id2": "9197"},
            {"sku_id": "lightbox", "category_id2": "13893"},
        ]
        electric = [
            {"sku_id": "accessory", "category_id2": "27511"},
            {"sku_id": "vehicle", "category_id2": "27509"},
        ]

        medical_allowed, medical_excluded = _prefilter_candidates("医疗器械", medical)
        electric_allowed, electric_excluded = _prefilter_candidates("电动车", electric)

        self.assertEqual(
            [item["sku_id"] for item in medical_allowed],
            ["device", "lightbox"],
        )
        self.assertEqual(medical_excluded, {"food": "平台类目ID与医疗器械不匹配"})
        self.assertEqual([item["sku_id"] for item in electric_allowed], ["vehicle"])
        self.assertEqual(electric_excluded, {"accessory": "平台类目ID与电动车不匹配"})

    def test_candidate_pool_keeps_up_to_thirty(self):
        from product_selection_agent.selector import build_candidate_pool

        goods = [
            {
                "skuId": str(index),
                "wname": f"商品{index}",
                "displayPrice": str(100 + index),
                "tab_category": "测试类目",
                "source_rank": index,
            }
            for index in range(1, 36)
        ]
        items = parse_all({"test": {"name": "测试来源", "goods": goods}})

        pool = build_candidate_pool(items, max_candidates=30)
        candidates = pool["测试来源"]["测试类目"]

        self.assertEqual(len(candidates), 30)
        self.assertEqual([item["candidate_rank"] for item in candidates], list(range(1, 31)))
        self.assertTrue(all("rank" not in item for item in candidates))

    def test_ai_rate_limit_uses_rolling_window(self):
        clock = [0.0]

        def fake_sleep(seconds):
            clock[0] += seconds

        with recommender._AI_RATE_LOCK:
            recommender._AI_REQUEST_TIMES.clear()
        try:
            with mock.patch.object(recommender.config, "AI_RPS_LIMIT", 2), \
                    mock.patch("product_selection_agent.recommender.time.monotonic", side_effect=lambda: clock[0]), \
                    mock.patch("product_selection_agent.recommender.time.sleep", side_effect=fake_sleep):
                recommender._wait_for_ai_rate_slot()
                recommender._wait_for_ai_rate_slot()
                recommender._wait_for_ai_rate_slot()
            self.assertGreaterEqual(clock[0], 1.0)
        finally:
            with recommender._AI_RATE_LOCK:
                recommender._AI_REQUEST_TIMES.clear()

    def test_stream_content_is_assembled(self):
        response = [
            b'data: {"choices":[{"delta":{"role":"assistant"}}]}\n',
            b'data: {"choices":[{"delta":{"content":"{\\"items\\":"}}]}\n',
            b'data: {"choices":[{"delta":{"content":"[]}"}}]}\n',
            b'data: [DONE]\n',
        ]
        self.assertEqual(_read_stream_content(response), '{"items":[]}')

    def test_stream_content_accepts_cumulative_gateway_snapshots(self):
        response = [
            b'data: {"choices":[{"delta":{"content":"{\\"items\\":"}}]}\n',
            b'data: {"choices":[{"delta":{"content":"{\\"items\\":[]}"}}]}\n',
            b'data: [DONE]\n',
        ]
        self.assertEqual(_read_stream_content(response), '{"items":[]}')

    def test_multiple_json_objects_are_merged_by_sku(self):
        text = (
            '{"items":[{"sku_id":"1","reason":"理由1","copy":"文案1"}]}\n'
            '{"items":[{"sku_id":"2","reason":"理由2","copy":"文案2"}]}'
        )
        parsed = _extract_json_object(text)
        self.assertEqual([item["sku_id"] for item in parsed["items"]], ["1", "2"])

    def test_multiple_ai_selection_objects_are_merged(self):
        text = (
            '{"selected":[{"sku_id":"1","rank":1,"reason":"理由1","copy":"文案1"}],'
            '"rejected":[],"shortfall_reason":""}\n'
            '{"selected":[{"sku_id":"2","rank":2,"reason":"理由2","copy":"文案2"}],'
            '"rejected":[{"sku_id":"3","reason":"不相关"}],'
            '"shortfall_reason":"仅2个合格"}'
        )
        parsed = _extract_json_object(text)
        self.assertEqual(
            [item["sku_id"] for item in parsed["selected"]],
            ["1", "2"],
        )
        self.assertEqual(parsed["rejected"][0]["sku_id"], "3")
        self.assertEqual(parsed["shortfall_reason"], "仅2个合格")

    def test_legacy_items_response_is_coerced_to_selected_protocol(self):
        from product_selection_agent.recommender import _coerce_ai_selection_payload

        parsed = {
            "items": [
                {"sku_id": "2", "reason": "理由2", "copy": "文案2"},
                {"sku_id": "1", "reason": "理由1", "copy": "文案1"},
            ]
        }

        coerced = _coerce_ai_selection_payload(parsed)

        self.assertEqual(
            [item["sku_id"] for item in coerced["selected"]],
            ["2", "1"],
        )
        self.assertEqual([item["rank"] for item in coerced["selected"]], [1, 2])
        self.assertTrue(
            any("兼容旧 items 协议" in item for item in coerced["_protocol_warnings"])
        )

    def test_top_level_array_is_coerced_to_selected_protocol(self):
        from product_selection_agent.recommender import _coerce_ai_selection_payload

        parsed = _extract_json_object(
            '[{"sku_id":"3","reason":"理由3","copy":"文案3"}]'
        )
        coerced = _coerce_ai_selection_payload(parsed)

        self.assertEqual(coerced["selected"][0]["sku_id"], "3")
        self.assertEqual(coerced["selected"][0]["rank"], 1)

    def test_ai_selection_protocol_filters_and_orders_candidates(self):
        from product_selection_agent.recommender import _normalize_ai_selection

        candidates = [{"sku_id": str(index)} for index in range(1, 13)]
        parsed = {
            "selected": [
                {"sku_id": "2", "rank": 2, "reason": "理由2", "copy": "文案2"},
                {"sku_id": "1", "rank": 1, "reason": "理由1", "copy": "文案1"},
                {"sku_id": "1", "rank": 3, "reason": "重复", "copy": "重复"},
                {"sku_id": "999", "rank": 4, "reason": "未知", "copy": "未知"},
            ] + [
                {
                    "sku_id": str(index),
                    "rank": index,
                    "reason": f"理由{index}",
                    "copy": f"文案{index}",
                }
                for index in range(3, 13)
            ],
            "rejected": [{"sku_id": "12", "reason": "相关性较弱"}],
            "shortfall_reason": "",
        }

        normalized = _normalize_ai_selection(parsed, candidates, limit=10)

        self.assertEqual(
            [item["sku_id"] for item in normalized["selected"]],
            [str(index) for index in range(1, 11)],
        )
        self.assertEqual(
            [item["rank"] for item in normalized["selected"]],
            list(range(1, 11)),
        )
        self.assertEqual(normalized["rejected"]["12"], "相关性较弱")
        self.assertEqual(normalized["rejected"]["11"], "AI 未入选（未返回详细理由）")
        self.assertTrue(any("未知 SKU" in warning for warning in normalized["warnings"]))
        self.assertTrue(any("重复 SKU" in warning for warning in normalized["warnings"]))
        self.assertTrue(any("超过上限" in warning for warning in normalized["warnings"]))

    def test_full_ai_selection_ignores_contradictory_shortfall_reason(self):
        from product_selection_agent.recommender import _normalize_ai_selection

        candidates = [{"sku_id": str(index)} for index in range(1, 11)]
        parsed = {
            "selected": [
                {
                    "sku_id": str(index),
                    "rank": index,
                    "reason": f"理由{index}",
                    "copy": f"文案{index}",
                }
                for index in range(1, 11)
            ],
            "rejected": [],
            "shortfall_reason": "未达到10个商品上限",
        }

        normalized = _normalize_ai_selection(parsed, candidates, limit=10)

        self.assertEqual(normalized["shortfall_reason"], "")
        self.assertTrue(
            any("入选已达上限" in warning for warning in normalized["warnings"])
        )

    def test_prompt_echo_is_not_used_as_shortfall_reason(self):
        from product_selection_agent.recommender import _normalize_ai_selection

        candidates = [{"sku_id": "1"}]
        parsed = {
            "selected": [
                {"sku_id": "1", "rank": 1, "reason": "理由", "copy": "文案"}
            ],
            "rejected": [],
            "shortfall_reason": (
                "相关有效商品达到10个时必须选择10个；不足10个时选择全部合格商品"
            ),
        }

        normalized = _normalize_ai_selection(parsed, candidates, limit=10)

        self.assertEqual(normalized["shortfall_reason"], "")
        self.assertTrue(
            any("不足说明疑似复制提示词" in warning for warning in normalized["warnings"])
        )

    def test_page_order_breaks_missing_metric_ties(self):
        goods = [
            {
                "skuId": str(index),
                "wname": f"商品{index}",
                "displayPrice": str(10 + index),
                "tab_category": "全场精选",
                "source_rank": index,
                "category_source": "source_fallback_no_tab",
            }
            for index in range(1, 13)
        ]
        items = parse_all({"black_friday": {"name": "黑色星期五", "goods": goods}})
        selection = select_top(items)
        selected = selection["黑色星期五"]["全场精选"]
        self.assertEqual(len(selected), 10)
        self.assertEqual(selected[0]["sku_id"], "1")
        # 单元测试不应因开发机存在本地模型配置而发起真实网络请求。
        with mock.patch.multiple(
            "product_selection_agent.recommender.config",
            AI_API_URL="",
            AI_API_KEY="",
            AI_MODEL="",
        ):
            recommendation = recommend(selection)
        block = recommendation["黑色星期五"]["全场精选"]
        self.assertEqual(len(block["products"]), 10)
        top = block["top_pick"]
        self.assertEqual(top["recommendation_mode"], "explainable_scoring")
        self.assertIn("当前展示价", top["reason"])

    def test_model_timeout_falls_back_and_opens_network_circuit(self):
        goods = [
            {
                "skuId": f"{category}-{index}",
                "wname": f"商品{category}-{index}",
                "displayPrice": str(10 + index),
                "tab_category": category,
                "source_rank": index,
            }
            for category in ("类目A", "类目B", "类目C", "类目D")
            for index in range(1, 11)
        ]
        selection = select_top(parse_all({"test": {"name": "测试来源", "goods": goods}}))
        with mock.patch.multiple(
            "product_selection_agent.recommender.config",
            AI_API_URL="http://model.test/chat/completions",
            AI_API_KEY="key",
            AI_MODEL="model",
            AI_CATEGORY_WORKERS=1,
            AI_CIRCUIT_FAILURE_THRESHOLD=3,
        ), mock.patch(
            "product_selection_agent.recommender._llm_select_category",
            side_effect=TimeoutError("timed out"),
        ) as select:
            blocks = recommend(selection)["测试来源"]
        self.assertEqual(blocks["类目A"]["recommendation_mode"], "explainable_scoring")
        self.assertEqual(len(blocks["类目A"]["products"]), 10)
        self.assertIn("TimeoutError", blocks["类目A"]["ai_error"])
        self.assertNotIn("模型熔断", blocks["类目B"]["ai_error"])
        self.assertIn("模型熔断", blocks["类目D"]["ai_error"])
        self.assertEqual(select.call_count, 3)

    def test_json_format_errors_do_not_open_network_circuit(self):
        goods = [
            {
                "skuId": f"{category}-{index}",
                "wname": f"商品{category}-{index}",
                "displayPrice": str(10 + index),
                "tab_category": category,
                "source_rank": index,
            }
            for category in ("类目A", "类目B", "类目C", "类目D")
            for index in range(1, 11)
        ]
        selection = select_top(parse_all({"test": {"name": "测试来源", "goods": goods}}))
        error = json.JSONDecodeError("Extra data", "{}{}", 2)
        with mock.patch.multiple(
            "product_selection_agent.recommender.config",
            AI_API_URL="http://model.test/chat/completions",
            AI_API_KEY="key",
            AI_MODEL="model",
            AI_CATEGORY_WORKERS=1,
            AI_CIRCUIT_FAILURE_THRESHOLD=3,
        ), mock.patch("product_selection_agent.recommender._llm_select_category", side_effect=error) as select:
            blocks = recommend(selection)["测试来源"]

        self.assertEqual(select.call_count, 4)
        self.assertTrue(all("模型熔断" not in block["ai_error"] for block in blocks.values()))

    def test_ai_selects_seven_from_thirty_in_one_request(self):
        from product_selection_agent.selector import build_candidate_pool

        goods = [
            {
                "skuId": str(index),
                "wname": f"商品{index}",
                "displayPrice": str(10 + index),
                "tab_category": "测试类目",
                "source_rank": index,
            }
            for index in range(1, 31)
        ]
        pool = build_candidate_pool(
            parse_all({"test": {"name": "测试来源", "goods": goods}}),
            max_candidates=30,
        )
        calls = []

        def fake_select(_source, _category, candidates, limit, _context=None):
            calls.append([item["sku_id"] for item in candidates])
            selected_ids = ["9", "2", "7", "1", "4", "6", "3"]
            return {
                "selected": [
                    {
                        "sku_id": sku_id,
                        "rank": rank,
                        "reason": f"AI理由{sku_id}",
                        "copy": f"AI文案{sku_id}",
                    }
                    for rank, sku_id in enumerate(selected_ids, 1)
                ],
                "rejected": {
                    item["sku_id"]: "类目相关性不足"
                    for item in candidates
                    if item["sku_id"] not in selected_ids
                },
                "shortfall_reason": "仅 7 个候选符合类目与选品要求",
                "warnings": [],
            }

        with mock.patch.multiple(
            "product_selection_agent.recommender.config",
            AI_API_URL="http://model.test/chat/completions",
            AI_API_KEY="key",
            AI_MODEL="model",
        ), mock.patch("product_selection_agent.recommender._llm_select_category", side_effect=fake_select) as select:
            block = recommend(pool)["测试来源"]["测试类目"]

        self.assertEqual(select.call_count, 1)
        self.assertEqual(len(calls[0]), 30)
        self.assertEqual(
            [item["sku_id"] for item in block["products"]],
            ["9", "2", "7", "1", "4", "6", "3"],
        )
        self.assertEqual([item["rank"] for item in block["products"]], list(range(1, 8)))
        self.assertEqual(block["shortfall_reason"], "仅 7 个候选符合类目与选品要求")
        self.assertEqual(block["recommendation_mode"], "llm_enhanced")
        self.assertEqual(block["candidate_decisions"]["9"]["ai_rank"], 1)
        self.assertEqual(
            block["candidate_decisions"]["30"]["rejection_reason"],
            "类目相关性不足",
        )

    def test_platform_category_prefilter_is_applied_before_ai_and_audited(self):
        candidates = [
            {
                "sku_id": "food",
                "name": "黑芝麻丸",
                "category_id2": "9195",
                "display_price": 20,
                "sales_num": 100000,
                "source_rank": 1,
                "discount_ratio": 0.2,
            },
            {
                "sku_id": "device",
                "name": "医用血压计",
                "category_id2": "9197",
                "display_price": 99,
                "sales_num": 1000,
                "source_rank": 2,
                "discount_ratio": 0.1,
            },
        ]

        def fake_select(_source, _category, ai_candidates, _limit, _context=None):
            self.assertEqual([item["sku_id"] for item in ai_candidates], ["device"])
            return {
                "selected": [
                    {
                        "sku_id": "device",
                        "rank": 1,
                        "reason": "医疗器械主体商品",
                        "copy": "血压计推荐",
                    }
                ],
                "rejected": {},
                "shortfall_reason": "仅1个相关有效商品",
                "warnings": [],
            }

        with mock.patch.multiple(
            "product_selection_agent.recommender.config",
            AI_API_URL="http://model.test/chat/completions",
            AI_API_KEY="key",
            AI_MODEL="model",
        ), mock.patch("product_selection_agent.recommender._llm_select_category", side_effect=fake_select):
            block = recommend({"国家补贴": {"医疗器械": candidates}})["国家补贴"]["医疗器械"]

        self.assertEqual([item["sku_id"] for item in block["products"]], ["device"])
        self.assertEqual(
            block["candidate_decisions"]["food"]["rejection_reason"],
            "平台类目ID与医疗器械不匹配",
        )

    def test_main_separates_candidate_pool_from_final_selection(self):
        from product_selection_agent import service as main

        goods = [
            {
                "skuId": str(index),
                "wname": f"商品{index}",
                "displayPrice": str(10 + index),
                "tab_category": "测试类目",
                "source_rank": index,
            }
            for index in range(1, 13)
        ]
        raw = {
            "test": {
                "name": "测试来源",
                "adapter": "offline",
                "status": "ok",
                "goods": goods,
            }
        }

        def fake_select(_source, _category, candidates, _limit, _context=None):
            selected = candidates[:7]
            return {
                "selected": [
                    {
                        "sku_id": item["sku_id"],
                        "rank": rank,
                        "reason": "AI理由",
                        "copy": "AI文案",
                    }
                    for rank, item in enumerate(selected, 1)
                ],
                "rejected": {
                    item["sku_id"]: "类目相关性不足"
                    for item in candidates[7:]
                },
                "shortfall_reason": "仅 7 个候选符合类目与选品要求",
                "warnings": [],
            }

        with mock.patch("product_selection_agent.service._load_offline", return_value=raw), mock.patch.multiple(
            "product_selection_agent.recommender.config",
            AI_API_URL="http://model.test/chat/completions",
            AI_API_KEY="key",
            AI_MODEL="model",
        ), mock.patch("product_selection_agent.recommender._llm_select_category", side_effect=fake_select):
            payload = main.run_selection(offline_path="ignored.json")

        candidates = payload["candidate_pool"]["测试来源"]["测试类目"]
        selected = payload["selection"]["测试来源"]["测试类目"]
        diagnostics = payload["diagnostics"]["sources"]["test"]
        self.assertEqual(len(candidates), 12)
        self.assertEqual(len(selected), 7)
        self.assertEqual(sum(bool(item["final_selected"]) for item in candidates), 7)
        self.assertEqual(sum(bool(item["ai_selected"]) for item in candidates), 7)
        self.assertIn("reco_score", candidates[0])
        self.assertIn("score_detail", candidates[0])
        self.assertEqual(diagnostics["candidate_goods_count"], 12)
        self.assertEqual(diagnostics["selected_goods_count"], 7)
        self.assertEqual(
            diagnostics["short_categories"]["测试类目"]["reason"],
            "仅 7 个候选符合类目与选品要求",
        )
        self.assertEqual(
            payload["contract"],
            "每来源 × 页面类目：AI 从最多30个候选中筛选最多Top10；合格不足不补齐",
        )
        self.assertTrue(payload["diagnostics"]["fetch_complete"])
        self.assertTrue(payload["diagnostics"]["ai_complete"])
        self.assertEqual(payload["diagnostics"]["ai_failed_categories"], [])
        self.assertTrue(payload["ai_complete"])

    def test_ai_completeness_lists_fallback_categories(self):
        from product_selection_agent import service as main

        recommendation = {
            "测试来源": {
                "成功类目": {
                    "recommendation_mode": "llm_enhanced",
                    "ai_error": "",
                },
                "回退类目": {
                    "recommendation_mode": "explainable_scoring",
                    "ai_error": "ValueError: AI 响应缺少 selected 数组",
                },
            }
        }

        diagnostics = main._ai_diagnostics(recommendation)

        self.assertFalse(diagnostics["ai_complete"])
        self.assertEqual(
            diagnostics["ai_failed_categories"],
            [
                {
                    "source": "测试来源",
                    "category": "回退类目",
                    "error": "ValueError: AI 响应缺少 selected 数组",
                }
            ],
        )

    def test_excel_contains_auditable_candidate_pool_sheet(self):
        from product_selection_agent import service as main
        from openpyxl import load_workbook

        product = {
            "source_name": "测试来源",
            "category_name": "测试类目",
            "candidate_rank": 1,
            "rank": 1,
            "name": "测试商品",
            "sku_id": "1",
            "display_price": 19.9,
            "sales_text": "已售100+",
            "sales_num": 100,
            "ai_selected": True,
            "ai_rank": 1,
            "rejection_reason": "",
            "shortfall_reason": "仅 1 个候选符合要求",
            "recommendation_mode": "llm_enhanced",
            "reco_score": 0.9,
            "score_detail": {"sales": 1.0},
            "reason": "AI理由",
            "copy": "AI文案",
        }
        payload = {
            "candidate_pool": {"测试来源": {"测试类目": [product]}},
            "selection": {"测试来源": {"测试类目": [product]}},
            "recommendation": {
                "测试来源": {
                    "测试类目": {
                        "products": [product],
                        "shortfall_reason": "仅 1 个候选符合要求",
                    }
                }
            },
            "diagnostics": {
                "sources": {
                    "test": {
                        "name": "测试来源",
                        "adapter": "offline",
                        "status": "ok",
                        "raw_goods_count": 1,
                        "candidate_goods_count": 1,
                        "category_count": 1,
                        "selected_goods_count": 1,
                        "short_categories": {
                            "测试类目": {
                                "selected_count": 1,
                                "reason": "仅 1 个候选符合要求",
                            }
                        },
                        "error": "",
                    }
                }
            },
        }

        with tempfile.TemporaryDirectory() as directory:
            path = main.save_excel(payload, directory, "test")
            workbook = load_workbook(path, read_only=True)
            self.assertEqual(
                workbook.sheetnames,
                ["推荐结果", "选品明细", "候选池", "运行诊断"],
            )
            recommendation_headers = [
                cell.value for cell in next(workbook["推荐结果"].iter_rows())
            ]
            self.assertNotIn("推荐模式", recommendation_headers)
            headers = [cell.value for cell in next(workbook["候选池"].iter_rows())]
            self.assertIn("候选排名", headers)
            self.assertIn("最终是否入选", headers)
            self.assertIn("AI是否入选", headers)
            self.assertIn("AI推荐顺位", headers)
            self.assertIn("淘汰原因", headers)
            self.assertIn("合格不足说明", headers)
            self.assertIn("选品方式", headers)
            self.assertNotIn("推荐模式", headers)
            mode_column = headers.index("选品方式") + 1
            self.assertEqual(
                workbook["候选池"].cell(row=2, column=mode_column).value,
                "智能推荐",
            )

    def test_fewer_selected_products_do_not_trigger_retry(self):
        goods = [
            {
                "skuId": str(index),
                "wname": f"商品{index}",
                "displayPrice": str(10 + index),
                "tab_category": "测试类目",
                "source_rank": index,
            }
            for index in range(1, 11)
        ]
        selection = select_top(parse_all({"test": {"name": "测试来源", "goods": goods}}))
        def fake_select(_source, _category, candidates, _limit, _context=None):
            selected = candidates[:-1]
            return {
                "selected": [
                    {
                        "sku_id": item["sku_id"],
                        "rank": rank,
                        "reason": "AI理由",
                        "copy": "AI文案",
                    }
                    for rank, item in enumerate(selected, 1)
                ],
                "rejected": {candidates[-1]["sku_id"]: "相关性不足"},
                "shortfall_reason": "仅 9 个候选符合要求",
                "warnings": [],
            }

        with mock.patch.multiple(
            "product_selection_agent.recommender.config",
            AI_API_URL="http://model.test/chat/completions",
            AI_API_KEY="key",
            AI_MODEL="model",
        ), mock.patch("product_selection_agent.recommender._llm_select_category", side_effect=fake_select) as select:
            block = recommend(selection)["测试来源"]["测试类目"]

        self.assertEqual(select.call_count, 1)
        self.assertEqual(len(block["products"]), 9)
        self.assertEqual(block["recommendation_mode"], "llm_enhanced")
        self.assertEqual(block["ai_error"], "")
        self.assertTrue(all(item["recommendation_mode"] == "llm_enhanced" for item in block["products"]))

    def test_categories_are_recommended_concurrently_and_keep_order(self):
        goods = [
            {
                "skuId": f"{category}-{index}",
                "wname": f"商品{category}-{index}",
                "displayPrice": str(10 + index),
                "tab_category": category,
                "source_rank": index,
            }
            for category in ("类目A", "类目B", "类目C", "类目D", "类目E")
            for index in range(1, 11)
        ]
        selection = select_top(parse_all({"test": {"name": "测试来源", "goods": goods}}))

        def fake_select(_source, _category, candidates, _limit, _context=None):
            time.sleep(0.2)
            return {
                "selected": [
                    {
                        "sku_id": item["sku_id"],
                        "rank": rank,
                        "reason": "AI理由",
                        "copy": "AI文案",
                    }
                    for rank, item in enumerate(candidates, 1)
                ],
                "rejected": {},
                "shortfall_reason": "",
                "warnings": [],
            }

        start = time.monotonic()
        with mock.patch.multiple(
            "product_selection_agent.recommender.config",
            AI_API_URL="http://model.test/chat/completions",
            AI_API_KEY="key",
            AI_MODEL="model",
            AI_CATEGORY_WORKERS=5,
            AI_RPS_LIMIT=5,
        ), mock.patch("product_selection_agent.recommender._llm_select_category", side_effect=fake_select) as select:
            result = recommend(selection)["测试来源"]
        elapsed = time.monotonic() - start

        self.assertLess(elapsed, 0.6)
        self.assertEqual(select.call_count, 5)
        self.assertEqual(list(result), list(selection["测试来源"]))
        self.assertTrue(all(block["recommendation_mode"] == "llm_enhanced" for block in result.values()))


if __name__ == "__main__":
    unittest.main()
