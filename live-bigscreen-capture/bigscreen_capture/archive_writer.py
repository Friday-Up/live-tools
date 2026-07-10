from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook
from openpyxl.drawing.image import Image as WorkbookImage
from openpyxl.styles import Alignment, Font, PatternFill

from .capture_manifest import CAPTURE_STEPS
from .file_naming import zip_filename


MANIFEST_NAME = "截图清单.xlsx"
HEADERS = ["计划整点", "实际执行时间", "直播间ID", "序号", "截图项", "文件名", "状态", "失败原因"]
SUMMARY_SHEET_NAME = "截图结果"
DETAIL_SHEET_NAME = "截图清单"
IMAGE_DISPLAY_MAX_WIDTH = 520
IMAGE_DISPLAY_MAX_HEIGHT = 330
SUMMARY_HEADERS = ["时间"] + ["%s %s" % (step.code, step.filename_label) for step in CAPTURE_STEPS]
SUMMARY_COLUMN_BY_STEP_CODE = {
    step.code: index + 2
    for index, step in enumerate(CAPTURE_STEPS)
}


def write_manifest_workbook(output_dir, records):
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / MANIFEST_NAME
    workbook = Workbook()
    try:
        summary_sheet = workbook.active
        summary_sheet.title = SUMMARY_SHEET_NAME
        detail_sheet = workbook.create_sheet(DETAIL_SHEET_NAME)

        _write_summary_sheet(summary_sheet, records)
        _write_detail_sheet(detail_sheet, records)
        workbook.save(path)
        return path
    finally:
        workbook.close()


def _write_summary_sheet(sheet, records):
    sheet.append(SUMMARY_HEADERS)
    _style_header_row(sheet, len(SUMMARY_HEADERS))
    sheet.freeze_panes = "B2"
    sheet.column_dimensions["A"].width = 14
    for column_index in range(2, len(SUMMARY_HEADERS) + 1):
        column_letter = sheet.cell(row=1, column=column_index).column_letter
        sheet.column_dimensions[column_letter].width = 75

    row_by_slot = {}
    for record in records:
        if record.planned_slot not in row_by_slot:
            row_index = sheet.max_row + 1
            row_by_slot[record.planned_slot] = row_index
            sheet.cell(row=row_index, column=1, value=record.planned_slot)
            sheet.row_dimensions[row_index].height = 255

    for record in records:
        row_index = row_by_slot[record.planned_slot]
        column_index = SUMMARY_COLUMN_BY_STEP_CODE.get(record.step_code)
        if not column_index:
            continue

        cell = sheet.cell(row=row_index, column=column_index)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        image_path = record.path if record.path else None
        if record.status == "成功" and image_path and image_path.is_file():
            try:
                image = WorkbookImage(str(image_path))
                _scale_image_for_display(image)
                image.anchor = cell.coordinate
                sheet.add_image(image)
            except Exception:
                cell.value = "成功（图片读取失败）"
        else:
            cell.value = _summary_status_text(record)


def _write_detail_sheet(sheet, records):
    sheet.append(HEADERS)
    _style_header_row(sheet, len(HEADERS))
    sheet.freeze_panes = "A2"
    widths = [14, 20, 14, 8, 24, 58, 10, 48]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = width

    for record in records:
        sheet.append(
            [
                record.planned_slot,
                record.executed_at.strftime("%Y-%m-%d %H:%M:%S") if record.executed_at else "",
                record.room_id,
                record.step_code,
                record.step_name,
                record.filename,
                record.status,
                record.error,
            ]
        )


def _style_header_row(sheet, column_count):
    header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for column_index in range(1, column_count + 1):
        cell = sheet.cell(row=1, column=column_index)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")


def _scale_image_for_display(image):
    scale = min(
        IMAGE_DISPLAY_MAX_WIDTH / image.width,
        IMAGE_DISPLAY_MAX_HEIGHT / image.height,
        1,
    )
    image.width = int(image.width * scale)
    image.height = int(image.height * scale)


def _summary_status_text(record):
    if record.status == "成功":
        return "成功（截图文件不存在）"
    if record.error:
        return "%s：%s" % (record.status, record.error)
    return record.status


def write_zip_archive(output_dir, room_id, captured_at):
    archive_path = output_dir / zip_filename(room_id, captured_at)
    manifest_path = output_dir / MANIFEST_NAME
    with ZipFile(archive_path, "w", ZIP_DEFLATED) as zf:
        if manifest_path.exists():
            zf.write(manifest_path, MANIFEST_NAME)
        for path in sorted(output_dir.rglob("*.png")):
            zf.write(path, path.relative_to(output_dir).as_posix())
    return archive_path
