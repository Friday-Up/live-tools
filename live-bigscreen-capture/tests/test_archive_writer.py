import tempfile
import unittest
import zipfile
from io import BytesIO
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image as PILImage

from bigscreen_capture.archive_writer import write_manifest_workbook, write_zip_archive
from bigscreen_capture.models import CaptureRecord


class ArchiveWriterTest(unittest.TestCase):
    def test_writes_summary_sheet_with_original_screenshots_scaled_for_display_and_detail_sheet(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            output_dir = Path(tmpdir)
            image_path = output_dir / "20260708_190000" / "蓝屏数据截图_46794566__20260708_190000_01_概览总览.png"
            gmv_path = output_dir / "20260708_190000" / "蓝屏数据截图_46794566__20260708_190000_15_GMVTop10.png"
            image_path.parent.mkdir(parents=True)
            PILImage.new("RGB", (800, 500), color="navy").save(image_path)
            PILImage.new("RGB", (800, 500), color="green").save(gmv_path)
            records = [
                CaptureRecord(
                    planned_slot="19:00",
                    executed_at=datetime(2026, 7, 8, 19, 0, 1),
                    room_id="46794566",
                    step_code="01",
                    step_name="概览总览",
                    filename=image_path.name,
                    status="成功",
                    path=image_path,
                ),
                CaptureRecord(
                    planned_slot="19:00",
                    executed_at=datetime(2026, 7, 8, 19, 0, 2),
                    room_id="46794566",
                    step_code="13",
                    step_name="用户画像_成交用户",
                    filename="蓝屏数据截图_46794566__20260708_190000_13_用户画像_成交用户.png",
                    status="失败",
                    error="下拉选择失败",
                ),
                CaptureRecord(
                    planned_slot="19:00",
                    executed_at=datetime(2026, 7, 8, 19, 0, 3),
                    room_id="46794566",
                    step_code="15",
                    step_name="GMVTop10",
                    filename=gmv_path.name,
                    status="成功",
                    path=gmv_path,
                ),
            ]

            manifest = write_manifest_workbook(output_dir, records)
            archive = write_zip_archive(
                output_dir,
                room_id="46794566",
                captured_at=datetime(2026, 7, 8, 19, 0, 0),
            )

            workbook = load_workbook(manifest)
            try:
                self.assertEqual(workbook.sheetnames, ["截图结果", "截图清单"])
                summary_sheet = workbook["截图结果"]
                detail_sheet = workbook["截图清单"]

                self.assertEqual(summary_sheet["A1"].value, "时间")
                self.assertEqual(summary_sheet["B1"].value, "01 概览总览")
                self.assertEqual(summary_sheet["P1"].value, "15 GMVTop10")
                self.assertEqual(summary_sheet["A2"].value, "19:00")
                self.assertIn("失败：下拉选择失败", summary_sheet["N2"].value)
                image_cells = {
                    (image.anchor._from.row + 1, image.anchor._from.col + 1)
                    for image in summary_sheet._images
                }
                self.assertEqual(image_cells, {(2, 2), (2, 16)})
                self.assertEqual(detail_sheet["A1"].value, "计划整点")
                self.assertEqual(detail_sheet["E2"].value, "概览总览")
                self.assertEqual(detail_sheet["E3"].value, "用户画像_成交用户")
            finally:
                workbook.close()
            with zipfile.ZipFile(manifest) as xlsx:
                media_dimensions = []
                drawing_xml = ""
                for name in xlsx.namelist():
                    if name.startswith("xl/media/"):
                        with PILImage.open(BytesIO(xlsx.read(name))) as image:
                            media_dimensions.append(image.size)
                    if name.startswith("xl/drawings/") and name.endswith(".xml"):
                        drawing_xml += xlsx.read(name).decode("utf-8")
            self.assertEqual(media_dimensions, [(800, 500), (800, 500)])
            self.assertIn('cx="4953000"', drawing_xml)
            self.assertIn('cy="3095625"', drawing_xml)

            with zipfile.ZipFile(archive) as zf:
                names = zf.namelist()
            self.assertIn("截图清单.xlsx", names)
            self.assertIn("20260708_190000/" + image_path.name, names)
            self.assertIn("20260708_190000/" + gmv_path.name, names)


if __name__ == "__main__":
    unittest.main()
