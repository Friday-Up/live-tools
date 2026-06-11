import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from utils.excel_handler import read_sku_list, write_results


def make_workbook(headers, rows):
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(tmp.name)
    return Path(tmp.name)


class ExcelHandlerTests(unittest.TestCase):
    def test_reads_required_sku_header_with_suffix(self):
        path = make_workbook(
            ["提交时间（自动）", "商品SKU（必填）", "提交者（自动）"],
            [["2026-06-11", "100264886683", "张三"]],
        )

        self.assertEqual(read_sku_list(str(path), "商品SKU"), [(2, "100264886683")])

    def test_rejects_missing_sku_header_instead_of_guessing_column_b(self):
        path = make_workbook(
            ["提交时间", "备注"],
            [["2026-06-11", "这不是SKU"]],
        )

        with self.assertRaisesRegex(ValueError, "未找到SKU列"):
            read_sku_list(str(path), "商品SKU")

    def test_write_results_adds_missing_output_headers(self):
        path = make_workbook(
            ["提交时间（自动）", "商品SKU（必填）", "提交者（自动）"],
            [["2026-06-11", "100264886683", "张三"]],
        )
        out_dir = tempfile.mkdtemp()

        output = write_results(
            file_path=str(path),
            results=[
                {
                    "row_index": 2,
                    "sku": "100264886683",
                    "price": 5.5,
                    "screenshot_path": None,
                    "status": "success",
                    "message": "ok",
                }
            ],
            threshold_price=6,
            output_dir=out_dir,
        )

        from openpyxl import load_workbook

        ws = load_workbook(output).active
        self.assertEqual(
            [ws.cell(1, i).value for i in range(1, 7)],
            ["提交时间（自动）", "商品SKU（必填）", "提交者（自动）", "价格", "图片", "备注"],
        )
        self.assertEqual(ws.cell(2, 4).value, 5.5)
        self.assertEqual(ws.cell(2, 6).value, "不符合上菜")

    def test_write_results_appends_output_headers_when_fallback_columns_are_occupied(self):
        path = make_workbook(
            ["提交时间", "商品SKU", "提交者", "活动名称"],
            [["2026-06-11", "100264886683", "张三", "618直播"]],
        )
        out_dir = tempfile.mkdtemp()

        output = write_results(
            file_path=str(path),
            results=[
                {
                    "row_index": 2,
                    "sku": "100264886683",
                    "price": 5.5,
                    "screenshot_path": None,
                    "status": "success",
                    "message": "ok",
                }
            ],
            threshold_price=6,
            output_dir=out_dir,
        )

        from openpyxl import load_workbook

        ws = load_workbook(output).active
        self.assertEqual(ws.cell(1, 4).value, "活动名称")
        self.assertEqual([ws.cell(1, i).value for i in range(5, 8)], ["价格", "图片", "备注"])
        self.assertEqual(ws.cell(2, 5).value, 5.5)
        self.assertEqual(ws.cell(2, 7).value, "不符合上菜")


if __name__ == "__main__":
    unittest.main()
